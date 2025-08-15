"""Export module: Contains all data export functionality.

This module handles Excel and XML export with full formatting and analysis data.
"""
import os
import re
import logging
from datetime import datetime
import xml.etree.ElementTree as ET
import xml.dom.minidom
from typing import Any

logger = logging.getLogger(__name__)


def export_excel(extractor: Any, output_path: str) -> None:
    """
    Export extracted transactions, account info, and analysis results to an Excel file. 
    """ 
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        logger.error("The 'openpyxl' library is required for Excel export. Please install it (pip install openpyxl).") 
        raise ImportError("openpyxl library required for Excel export")

    try:
        wb = openpyxl.Workbook() 
        # Remove default sheet if needed, or rename it 
        if "Sheet" in wb.sheetnames:
            default_sheet = wb["Sheet"]
            wb.remove(default_sheet)

        header_font = Font(bold=True) 

        # --- 1. All Transactions Sheet --- 
        all_trans_ws = wb.create_sheet("All Transactions")
        all_trans_ws.column_dimensions['A'].width = 18 # Date (DD-Month-YYYY) 
        all_trans_ws.column_dimensions['B'].width = 50 # Description
        all_trans_ws.column_dimensions['C'].width = 15 # Debit
        all_trans_ws.column_dimensions['D'].width = 15 # Credit 
        all_trans_ws.column_dimensions['E'].width = 15 # Balance 

        # Add bank name, account number, and transaction ID at the top 
        current_header_row = 1
        if extractor.account_info.get('bank_name'):
            all_trans_ws.cell(row=current_header_row, column=1, value="Bank Name:").font = header_font
            all_trans_ws.cell(row=current_header_row, column=2, value=extractor.account_info['bank_name'])
            current_header_row += 1 

        # Add account number prominently 
        if extractor.account_info.get('account_number'):
            all_trans_ws.cell(row=current_header_row, column=1, value="Account Number:").font = header_font
            cell_acc_num = all_trans_ws.cell(row=current_header_row, column=2, value=extractor.account_info['account_number'])
            cell_acc_num.font = Font(bold=True, size=12) # Make account number bold and slightly larger 
            current_header_row += 1

        all_trans_ws.cell(row=current_header_row, column=1, value="Transaction ID:").font = header_font 
        all_trans_ws.cell(row=current_header_row, column=2, value=extractor.unique_id)
        current_header_row += 2 # Add a blank row before headers

        headers_all = ['Date', 'Description', 'Category', 'Debit', 'Credit', 'Balance'] 
        for col, header in enumerate(headers_all, 1):
            cell = all_trans_ws.cell(row=current_header_row, column=col)
            cell.value = header
            cell.font = header_font 

        # Start data rows after the headers
        data_start_row = current_header_row + 1
        for row_idx, trans in enumerate(extractor.transactions, data_start_row):
            formatted_date = _format_date_for_excel(trans.get('date')) 
            amount = trans.get('amount', 0) 
            debit_val = abs(amount) if amount is not None and amount < 0 else None
            credit_val = amount if amount is not None and amount > 0 else None
            balance_val = trans.get('balance')
            category_val = trans.get('category', 'Other')
            all_trans_ws.cell(row=row_idx, column=1, value=formatted_date) 
            all_trans_ws.cell(row=row_idx, column=2, value=trans.get('description'))
            all_trans_ws.cell(row=row_idx, column=3, value=category_val)
            cell_debit = all_trans_ws.cell(row=row_idx, column=4, value=debit_val)
            cell_credit = all_trans_ws.cell(row=row_idx, column=5, value=credit_val)
            cell_balance = all_trans_ws.cell(row=row_idx, column=6, value=balance_val)
            if debit_val is not None: cell_debit.number_format = '#,##0.00'
            if credit_val is not None: cell_credit.number_format = '#,##0.00'
            if balance_val is not None: cell_balance.number_format = '#,##0.00'


        # --- 2. High Value Transactions Sheet --- 
        high_value_ws = wb.create_sheet("High Value Transactions") 
        high_value_ws.column_dimensions['A'].width = 18 # Date 
        high_value_ws.column_dimensions['B'].width = 50 # Description
        high_value_ws.column_dimensions['C'].width = 15 # Debit
        high_value_ws.column_dimensions['D'].width = 15 # Credit
        high_value_ws.column_dimensions['E'].width = 15 # Balance

        headers_high = ['Date', 'Description', 'Category', 'Debit', 'Credit', 'Balance'] 
        for col, header in enumerate(headers_high, 1): 
            cell = high_value_ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font

        for row_idx, trans in enumerate(extractor.high_value_transactions, 2): # Use extractor.high_value_transactions 
             formatted_date = _format_date_for_excel(trans.get('date')) 
             amount = trans.get('amount', 0)
             debit_val = abs(amount) if amount is not None and amount < 0 else None
             credit_val = amount if amount is not None and amount > 0 else None
             balance_val = trans.get('balance')
             category_val = trans.get('category', 'Other')
             high_value_ws.cell(row=row_idx, column=1, value=formatted_date) 
             high_value_ws.cell(row=row_idx, column=2, value=trans.get('description'))
             high_value_ws.cell(row=row_idx, column=3, value=category_val)
             cell_debit_hv = high_value_ws.cell(row=row_idx, column=4, value=debit_val)
             cell_credit_hv = high_value_ws.cell(row=row_idx, column=5, value=credit_val)
             cell_balance_hv = high_value_ws.cell(row=row_idx, column=6, value=balance_val)
             if debit_val is not None: cell_debit_hv.number_format = '#,##0.00'
             if credit_val is not None: cell_credit_hv.number_format = '#,##0.00'
             if balance_val is not None: cell_balance_hv.number_format = '#,##0.00'


        # --- 3. Analysis Sheet --- 
        analysis_ws = wb.create_sheet("Analysis") 
        analysis_ws.column_dimensions['A'].width = 35
        analysis_ws.column_dimensions['B'].width = 25
        analysis_ws.column_dimensions['C'].width = 18
        analysis_ws.column_dimensions['D'].width = 18
        analysis_ws.column_dimensions['E'].width = 30 # Wider for date ranges

        current_row = 1 

        # --- Header Section ---
        analysis_ws.cell(row=current_row, column=1, value="Extraction ID:").font = header_font
        analysis_ws.cell(row=current_row, column=2, value=extractor.unique_id)
        current_row += 1
        analysis_ws.cell(row=current_row, column=1, value="Start Date:").font = header_font
        analysis_ws.cell(row=current_row, column=2, value=extractor.start_date.strftime('%d-%B-%Y'))
        current_row += 1
        # Always use the bank name as selected by the user (extractor.account_info['bank_name'])
        bank_name_val = extractor.account_info.get('bank_name', 'N/A')
        analysis_ws.cell(row=current_row, column=1, value="Bank Name:").font = header_font
        analysis_ws.cell(row=current_row, column=2, value=bank_name_val)
        current_row += 2
        # --- Account Information Section ---
        analysis_ws.cell(row=current_row, column=1, value="Account Information").font = header_font
        current_row += 1
        analysis_ws.cell(row=current_row, column=1, value="Account Number").font = header_font
        analysis_ws.cell(row=current_row, column=2, value=extractor.account_info.get('account_number', 'N/A'))
        current_row += 1
        if extractor.account_info.get('tax_id'):
            analysis_ws.cell(row=current_row, column=1, value="Tax ID").font = header_font
            analysis_ws.cell(row=current_row, column=2, value=extractor.account_info.get('tax_id'))
            current_row += 1
        # Only show one Opening Balance, and match it to the first month in Monthly Summary if available
        monthly_summary_data = extractor.analysis_results.get('monthly_summary', {})
        first_month = None
        first_month_ob = None
        if monthly_summary_data:
            first_month = sorted(monthly_summary_data.keys())[0]
            first_month_ob = monthly_summary_data[first_month].get('opening_balance')
        if first_month_ob is not None:
            try:
                ob_val = float(first_month_ob)
                analysis_ws.cell(row=current_row, column=1, value="Opening Balance").font = header_font
                analysis_ws.cell(row=current_row, column=2, value=ob_val)
                current_row += 1
            except Exception:
                pass
        if extractor.account_info.get('closing_balance') is not None:
            try:
                cb_val = float(extractor.account_info.get('closing_balance'))
                analysis_ws.cell(row=current_row, column=1, value="Closing Balance").font = header_font
                analysis_ws.cell(row=current_row, column=2, value=cb_val)
                current_row += 1
            except Exception:
                pass
        current_row += 1
        # --- Monthly Summary Section ---
        analysis_ws.cell(row=current_row, column=1, value="Monthly Summary").font = header_font
        current_row += 1
        headers_monthly = [
            "Month",
            "Opening Balance",
            "Balance After 5th",
            "Balance After 15th",
            "Balance After 25th"
        ]
        for col, header in enumerate(headers_monthly, 1):
            cell = analysis_ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
        current_row += 1
        monthly_summary_data = extractor.analysis_results.get('monthly_summary', {})
        if monthly_summary_data:
            for month, data in sorted(monthly_summary_data.items()):
                opening_bal = data.get('opening_balance')
                bal_after_5 = data.get('balance_after_5th')
                bal_after_15 = data.get('balance_after_15th')
                bal_after_25 = data.get('balance_after_25th')
                analysis_ws.cell(row=current_row, column=1, value=month)
                cell_ob = analysis_ws.cell(row=current_row, column=2, value=opening_bal)
                cell_b5 = analysis_ws.cell(row=current_row, column=3, value=bal_after_5)
                cell_b15 = analysis_ws.cell(row=current_row, column=4, value=bal_after_15)
                cell_b25 = analysis_ws.cell(row=current_row, column=5, value=bal_after_25)
                if opening_bal is not None: cell_ob.number_format = '#,##0.00'
                if bal_after_5 is not None: cell_b5.number_format = '#,##0.00'
                if bal_after_15 is not None: cell_b15.number_format = '#,##0.00'
                if bal_after_25 is not None: cell_b25.number_format = '#,##0.00'
                current_row += 1
        else:
            analysis_ws.cell(row=current_row, column=1, value="No monthly summary data available.")
            current_row += 1
        current_row += 1
        # --- Identified EMIs/Loans Section --- 
        analysis_ws.cell(row=current_row, column=1, value="Identified EMIs/Recurring Debits").font = Font(bold=True)
        current_row += 1
        emis_data = extractor.analysis_results.get('identified_emis', []) 
        if emis_data: 
            analysis_ws.cell(row=current_row, column=1, value="Description Pattern").font = Font(bold=True)
            analysis_ws.cell(row=current_row, column=2, value="Estimated Amount").font = Font(bold=True)
            analysis_ws.cell(row=current_row, column=3, value="Occurrences").font = Font(bold=True)
            analysis_ws.cell(row=current_row, column=4, value="First Date").font = Font(bold=True)
            analysis_ws.cell(row=current_row, column=5, value="Last Date").font = Font(bold=True) 
            current_row += 1
            for emi in emis_data:
                est_amount = emi.get('estimated_amount')
                first_date = _format_date_for_excel(emi.get('first_occurrence_date'))
                last_date = _format_date_for_excel(emi.get('last_occurrence_date')) 

                analysis_ws.cell(row=current_row, column=1, value=emi.get('description_pattern'))
                cell_emi_amt = analysis_ws.cell(row=current_row, column=2, value=est_amount)
                analysis_ws.cell(row=current_row, column=3, value=emi.get('occurrences'))
                analysis_ws.cell(row=current_row, column=4, value=first_date) 
                analysis_ws.cell(row=current_row, column=5, value=last_date)

                if est_amount is not None: cell_emi_amt.number_format = '#,##0.00'
                current_row += 1
        else:
             analysis_ws.cell(row=current_row, column=1, value="None Identified") 
             current_row += 1
        current_row += 1 # Add gap


        # --- Identified Salary Section --- 
        analysis_ws.cell(row=current_row, column=1, value="Identified Salary/Regular Credits").font = Font(bold=True)
        current_row += 1
        salary_data = extractor.analysis_results.get('identified_salary', []) 
        if salary_data: 
            analysis_ws.cell(row=current_row, column=1, value="Identification Type").font = Font(bold=True)
            analysis_ws.cell(row=current_row, column=2, value="Description Sample").font = Font(bold=True)
            analysis_ws.cell(row=current_row, column=3, value="Estimated Amount").font = Font(bold=True)
            analysis_ws.cell(row=current_row, column=4, value="Occurrences").font = Font(bold=True) 
            analysis_ws.cell(row=current_row, column=5, value="Date Range").font = Font(bold=True)
            current_row += 1
            for sal in salary_data:
                est_amount = sal.get('estimated_amount') 
                first_date = _format_date_for_excel(sal.get('first_occurrence_date')) 
                last_date = _format_date_for_excel(sal.get('last_occurrence_date'))

                analysis_ws.cell(row=current_row, column=1, value=sal.get('type'))
                analysis_ws.cell(row=current_row, column=2, value=sal.get('description_sample'))
                cell_sal_amt = analysis_ws.cell(row=current_row, column=3, value=est_amount)
                analysis_ws.cell(row=current_row, column=4, value=sal.get('occurrences')) 
                analysis_ws.cell(row=current_row, column=5, value=f"{first_date} to {last_date}") # Combined date range

                if est_amount is not None: cell_sal_amt.number_format = '#,##0.00'
                current_row += 1
        else: 
            analysis_ws.cell(row=current_row, column=1, value="None Identified")
            current_row += 1
        # End of Analysis Sheet content


        # --- Save Workbook --- 
        wb.save(output_path)
        logger.info(f"Excel file with analysis saved successfully at: {output_path}") 

    except Exception as e:
        logger.error(f"Error exporting data to Excel: {e}", exc_info=True) 
        raise


def export_xml(extractor: Any, output_path: str = None) -> str:
    """Export extracted data to a structured XML file.""" 
    if not output_path:
        base_name = os.path.splitext(os.path.basename(extractor.pdf_path))[0]
        output_path = f"{base_name}_{extractor.unique_id}_summary.xml"

    logger.info(f"Exporting data to XML: {output_path}")

    try:
        # --- Create Root Element --- 
        root = ET.Element('BankStatement') 
        root.set('transactionId', extractor.unique_id)
        root.set('sourcePdf', os.path.basename(extractor.pdf_path))
        root.set('generatedTimestamp', datetime.now().isoformat())
        if extractor.account_info.get('bank_name'):
            root.set('bankName', extractor.account_info['bank_name'])

        # --- Account Info Element --- 
        account_info_elem = ET.SubElement(root, 'AccountInformation') 
        for key, value in extractor.account_info.items():
            # Clean key for XML tag name (basic) 
            tag_name = re.sub(r'\s+', '_', key).strip()
            tag_name = re.sub(r'[^a-zA-Z0-9_.-]', '', tag_name) # Allow dot and hyphen
            # Ensure starts with letter or underscore
            if not tag_name or not (tag_name[0].isalpha() or tag_name[0] == '_'):
                tag_name = '_' + tag_name # Prepend underscore if invalid start

            if tag_name: # Ensure tag name is not empty after cleaning 
                ET.SubElement(account_info_elem, tag_name).text = str(value) if value is not None else ""

        # --- Account Summary Element --- 
        summary = _generate_account_summary(extractor)
        summary_elem = ET.SubElement(root, 'AccountSummary')
        for section, data in summary.items(): 
            section_tag = re.sub(r'\s+', '', section) # e.g., SummaryInfo, MonthwiseDetails
            section_elem = ET.SubElement(summary_elem, section_tag)

            if isinstance(data, dict):
                for key, value in data.items():
                    key_tag = re.sub(r'\s+', '', key) # e.g., NameOfTheAccountHolder
                    key_tag = re.sub(r'[^a-zA-Z0-9_.-]', '', key_tag)
                    if not key_tag or not (key_tag[0].isalpha() or key_tag[0] == '_'):
                       key_tag = '_' + key_tag

                    if isinstance(value, dict): # Handle nested dict like Monthwise Details
                        sub_section_elem = ET.SubElement(section_elem, key_tag)
                        for sub_key, sub_value in value.items():
                            sub_key_tag = re.sub(r'\s+', '', sub_key)
                            sub_key_tag = re.sub(r'[^a-zA-Z0-9_.-]', '', sub_key_tag)
                            if not sub_key_tag or not (sub_key_tag[0].isalpha() or sub_key_tag[0] == '_'):
                               sub_key_tag = '_' + sub_key_tag
                            ET.SubElement(sub_section_elem, sub_key_tag).text = str(sub_value) if sub_value is not None else ""
                    elif isinstance(value, list): # Handle lists like EMIs/Salary
                        list_elem = ET.SubElement(section_elem, key_tag)
                        for item in value:
                            if isinstance(item, dict):
                                item_tag = "Item" # Generic tag for list items
                                item_elem = ET.SubElement(list_elem, item_tag)
                                for item_key, item_val in item.items():
                                    item_key_tag = re.sub(r'\s+', '', item_key)
                                    item_key_tag = re.sub(r'[^a-zA-Z0-9_.-]', '', item_key_tag)
                                    if not item_key_tag or not (item_key_tag[0].isalpha() or item_key_tag[0] == '_'):
                                       item_key_tag = '_' + item_key_tag
                                    ET.SubElement(item_elem, item_key_tag).text = str(item_val) if item_val is not None else ""
                    else: # Simple key-value pair
                         ET.SubElement(section_elem, key_tag).text = str(value) if value is not None else ""


        # --- High Value Transactions Element --- 
        high_value_elem = ET.SubElement(root, 'HighValueTransactions')
        high_value_elem.set('threshold', str(extractor.high_value_threshold)) 
        if extractor.high_value_transactions:
             for tx in extractor.high_value_transactions:
                  tx_elem = ET.SubElement(high_value_elem, 'Transaction') 
                  tx_elem.set('id', tx.get('transaction_id', ''))
                  for key, value in tx.items():
                       if key == 'is_high_value': continue # Skip flag in XML details 
                       # Clean key for XML tag name (reuse logic) 
                       tag_name = re.sub(r'\s+', '_', key).strip()
                       tag_name = re.sub(r'[^a-zA-Z0-9_.-]', '', tag_name)
                       if not tag_name or not (tag_name[0].isalpha() or tag_name[0] == '_'):
                           tag_name = '_' + tag_name

                       if tag_name: 
                            # Format numbers 
                            if isinstance(value, (int, float)):
                                 text_value = f"{value:.2f}" 
                            elif isinstance(value, datetime):
                                 text_value = value.strftime('%Y-%m-%d') # Format date consistently
                            else: 
                                 text_value = str(value) if value is not None else ""
                            ET.SubElement(tx_elem, tag_name).text = text_value


        # --- All Transactions Element --- 
        transactions_elem = ET.SubElement(root, 'AllTransactions')
        if extractor.transactions:
             for tx in extractor.transactions:
                  tx_elem = ET.SubElement(transactions_elem, 'Transaction')
                  tx_elem.set('id', tx.get('transaction_id', '')) 
                  for key, value in tx.items():
                       if key == 'is_high_value': continue # Skip flag 
                       # Clean key for XML tag name (reuse logic) 
                       tag_name = re.sub(r'\s+', '_', key).strip()
                       tag_name = re.sub(r'[^a-zA-Z0-9_.-]', '', tag_name)
                       if not tag_name or not (tag_name[0].isalpha() or tag_name[0] == '_'):
                           tag_name = '_' + tag_name

                       if tag_name: 
                            # Format numbers and dates 
                            if isinstance(value, (int, float)): 
                                 text_value = f"{value:.2f}"
                            elif isinstance(value, datetime): 
                                 text_value = value.strftime('%Y-%m-%d')
                            else:
                                 text_value = str(value) if value is not None else "" 
                            ET.SubElement(tx_elem, tag_name).text = text_value

        # --- Convert ET to string with pretty printing --- 
        rough_string = ET.tostring(root, 'utf-8')
        reparsed = xml.dom.minidom.parseString(rough_string)
        pretty_xml_as_string = reparsed.toprettyxml(indent="  ", encoding='utf-8') # Ensure utf-8 output

        # --- Write to file --- 
        with open(output_path, "wb") as f: # Use 'wb' for bytes output from toprettyxml
             f.write(pretty_xml_as_string) 

        logger.info(f"Successfully exported data to {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"Failed to export data to XML: {e}", exc_info=extractor.debug) 
        raise # Re-raise the exception


def _format_date_for_excel(date_input):
    """Format date object or string into a more detailed format (DD-Month-YYYY).""" 
    if date_input is None:
        return None

    date_obj = None
    if isinstance(date_input, datetime):
        date_obj = date_input
    elif isinstance(date_input, str):
        date_str = date_input.strip()
        if not date_str or date_str.lower() in ['nan', 'none', 'date']: return None 
        try:
            # Try standard format first 
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            # Try other common formats 
            for fmt in [
                '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d',
                '%d/%m/%y', '%d-%m-%y', '%Y-%m-%d',
                '%m/%d/%Y', '%m-%d-%Y',
                '%d-%b-%Y', '%d %b %Y',  # For dates like 01-Jan-2024 or 01 Jan 2024 
                '%b %d %Y',              # For dates like Jan 01 2024
                '%d-%B-%Y', '%d %B %Y',  # For dates like 01-January-2024 or 01 January 2024 
                '%B %d %Y',               # For dates like January 01 2024 
                '%d %b, %Y',              # For dates like 01 Feb, 2024 
                '%d %B, %Y',              # For dates like 01 February, 2024 
                '%b %d, %Y',              # For dates like Feb 01, 2024
                '%B %d, %Y',              # For dates like February 01, 2024 
            ]:
                try:
                    parsed = datetime.strptime(date_str, fmt)
                    # Handle 2-digit years 
                    if parsed.year < 100:
                        if parsed.year < 50:
                            parsed = parsed.replace(year=parsed.year + 2000) 
                        else:
                            parsed = parsed.replace(year=parsed.year + 1900) 
                    date_obj = parsed
                    break # Exit loop on successful parse 
                except ValueError:
                    continue # Try next format 

            # Regex fallback if formats fail 
            if date_obj is None:
                try:
                    # Match patterns like DD-MM-YYYY, DD/MM/YYYY, etc. 
                    match = re.match(r'(\d{1,2})[-/.\s](\d{1,2})[-/.\s](\d{2,4})', date_str)
                    if match:
                        day, month, year = match.groups() 
                        day = int(day); month = int(month); year = int(year) 
                        if year < 100: # Handle 2-digit year 
                            year = year + 2000 if year < 50 else year + 1900
                        try: 
                            date_obj = datetime(year, month, day)
                        except ValueError: 
                            pass 
                except Exception: 
                    pass
    else: # Input is not datetime or string
        logger.warning(f"Invalid input type for date formatting: {type(date_input)}")
        return None

    # If we have a valid date object, format it
    if date_obj:
         try:
             return date_obj.strftime('%d-%B-%Y') # e.g., "01-January-2024" 
         except ValueError: # Handle potential issues with very old dates if necessary
             logger.warning(f"Could not format date object: {date_obj}")
             return str(date_input) # Return original input as string
    else:
        logger.warning(f"Failed to parse or format date: {date_input}")
        return str(date_input) # Return original input as string if all fails


def _generate_account_summary(extractor):
    """Generate a dictionary summarizing account activity.""" 
    import calendar
    
    summary = {} 

    # --- Summary Info Section --- 
    summary['Summary Info'] = {
        'Name of the Account Holder': extractor.account_info.get('owner_name', 'N/A'),
        'Address': extractor.account_info.get('address', 'N/A'), # Assuming you might add address extraction
        'Email': extractor.account_info.get('email', 'N/A'), # Assuming you might add email extraction
        'PAN': extractor.account_info.get('pan', 'N/A'), # Assuming you might add PAN extraction
        'Mobile Number': extractor.account_info.get('mobile', 'N/A'), 
        'Name of the Bank': extractor.account_info.get('bank_name', 'N/A'),
        'Account Number': extractor.account_info.get('account_number', 'N/A'),
        'Account Type': extractor.account_info.get('account_type', 'N/A'),
        'EMI Amount': 'N/A', # Placeholder, calculate from analysis if needed
        'Transaction ID': extractor.unique_id 
    }

    # --- Monthwise Details Section --- 
    # Use analysis_results['monthly_summary'] which has opening balance
    monthly_summary_data = extractor.analysis_results.get('monthly_summary', {})
    if monthly_summary_data: 
        monthly_details = {}
        all_balances = [] # For average calculation
        for month_key in sorted(monthly_summary_data.keys()): 
            try:
                data = monthly_summary_data[month_key]
                year, month = map(int, month_key.split('-')) 
                month_name = calendar.month_name[month]
                short_month = f"{month_name[:3]}-{str(year)[2:]}"  # e.g., "Jun-24" 

                # Get opening and closing balance for the month 
                # Use the balances calculated in _calculate_monthly_balances
                opening_balance = data.get('opening_balance') # Balance before first txn of month
                balance_on_5th = data.get('balance_after_5th') # Balance on/after 5th

                # Find closing balance for the month (last transaction's balance)
                closing_balance = None
                transactions_this_month = [
                    t['balance'] for t in extractor.transactions
                    if t.get('date') and t['date'].startswith(month_key) and t.get('balance') is not None
                ]
                if transactions_this_month:
                    closing_balance = transactions_this_month[-1]
                    all_balances.append(closing_balance) # Add month-end balance for averaging

                monthly_details[short_month] = { 
                    'Opening Balance': f"{opening_balance:.2f}" if opening_balance is not None else 'N/A',
                    'Closing Balance': f"{closing_balance:.2f}" if closing_balance is not None else 'N/A',
                    'Balance as on 5th': f"{balance_on_5th:.2f}" if balance_on_5th is not None else 'N/A' 
                }
            except Exception as e:
                logger.error(f"Error processing month {month_key} for summary: {e}") 
                continue 

        summary['Monthwise Details'] = monthly_details 

        # --- Calculate Averages (Simplified Example) --- 
        # Requires more transactions over longer periods for meaningful averages
        num_months = len(all_balances)
        avg_3m = avg_6m = avg_12m = 'NA'
        if num_months >= 3: avg_3m = f"{(sum(all_balances[-3:]) / 3):.2f}"
        if num_months >= 6: avg_6m = f"{(sum(all_balances[-6:]) / 6):.2f}"
        if num_months >= 12: avg_12m = f"{(sum(all_balances[-12:]) / 12):.2f}"

        # Calculate transaction averages (requires tracking monthly counts/amounts)
        # This part needs more detailed implementation within transaction processing or analysis
        avg_credit_count_6m = 'NA'
        avg_credit_amount_6m = 'NA'
        avg_debit_count_6m = 'NA'

    else:
        summary['Monthwise Details'] = {} 
        avg_3m = avg_6m = avg_12m = 'NA'
        avg_credit_count_6m = 'NA'
        avg_credit_amount_6m = 'NA'
        avg_debit_count_6m = 'NA'

    # --- Overall Summary Section --- 
    summary['Overall Summary'] = {
        'Average Balance of 3 Months': avg_3m,
        'Average Balance of 6 Months': avg_6m,
        'Average Balance of 12 Months': avg_12m,
        'Average of No.of Credit Transaction of 6 months': avg_credit_count_6m, 
        'Average of Amount of Credit Transaction of 6 Months': avg_credit_amount_6m,
        'Average No.of Debit Transaction of 6 months': avg_debit_count_6m 
    }

    # Add identified EMIs/Salary to summary if desired
    summary['Analysis'] = {
        'Identified EMIs': extractor.analysis_results.get('identified_emis', []),
        'Identified Salary': extractor.analysis_results.get('identified_salary', [])
    }

    return summary


