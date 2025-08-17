# --- bank_statement_extractor.py (Updated) ---

import os
import re
import pandas as pd
import xml.etree.ElementTree as ET
import xml.dom.minidom
import PyPDF2
import argparse
from datetime import datetime, timedelta # Added timedelta
from tabula.io import read_pdf
import logging
import calendar
import uuid
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font
import traceback
from fuzzywuzzy import fuzz
from collections import Counter
import pdfplumber
from typing import Dict, List, Optional, Tuple

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- Replace basicConfig with this more detailed setup ---
log_directory = "logs" # Choose a directory name
if not os.path.exists(log_directory):
    os.makedirs(log_directory)

log_filename = os.path.join(log_directory, 'extractor.log')

# Get the root logger and set its level to DEBUG to capture everything
root_logger = logging.getLogger()
root_logger.setLevel(logging.DEBUG) # Capture DEBUG and above

log_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# Console Handler (optional - keep INFO level for console)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO) # Set console level (e.g., INFO)
console_handler.setFormatter(log_formatter)
root_logger.addHandler(console_handler)

# File Handler (set level to DEBUG for file)
file_handler = logging.FileHandler(log_filename, mode='a') # 'a' for append
file_handler.setLevel(logging.DEBUG) # Log DEBUG level to file
file_handler.setFormatter(log_formatter)
root_logger.addHandler(file_handler)

# Get the specific logger for your module
logger = logging.getLogger('BankStatementExtractor')
# logger.propagate = False # Optional: Set to True (default) or False if you want to control if messages also go to root logger handlers

# --- End logging setup ---

# --- Template Registry ---
TEMPLATE_PARSERS = {
    "KOTAK_SINGLE_COL": "_parse_kotak_single_col",
    "HDFC_SEPARATE_COLS": "_parse_hdfc_separate_cols",
    "SBI_SEPARATE_COLS": "_parse_sbi_separate_cols",
    # Add more as needed
}

class BankStatementExtractor:
    # Update the __init__ method signature
    def __init__(self, pdf_path, start_date_str, end_date_str, password=None, debug=False, high_value_threshold=1000, bank_name=None):
        """Initialize the bank statement extractor.""" 
        self.pdf_path = pdf_path
        self.password = password
        self.debug = debug
        self.high_value_threshold = high_value_threshold
        self.unique_id = str(uuid.uuid4())[:8]
        self.transactions = []
        self.high_value_transactions = []
        self.account_info = {}
        self.analysis_results = {}
        self.monthly_balances = {}
        self.statement_opening_balance = None
        self.found_first_row_balance = False

        # Store the bank name provided by the user 
        if bank_name:
            self.account_info['bank_name'] = bank_name
            logger.info(f"Using provided bank name: {bank_name}")

        # Parse date range 
        try:
            self.start_date = datetime.strptime(start_date_str, '%m/%Y')
            self.end_date = datetime.strptime(end_date_str, '%m/%Y')
            # Set current parsing month/year to start date
            self.current_month = self.start_date.month
            self.current_year = self.start_date.year
        except ValueError as e:
            logger.error(f"Invalid date format. Expected MM/YYYY, got: {start_date_str} and {end_date_str}") 
            raise ValueError(f"Invalid date format. Please use MM/YYYY format (e.g., 01/2024)") from e

        # --- Add variable for first row balance ---
        self.statement_opening_balance = None # Store balance found from first transaction row 
        self.found_first_row_balance = False # Flag to ensure we only capture it once 
        # --- End add variable ---

        # --- Add Start/End Date Processing --- 
        self.start_year = None
        self.start_month = None
        self.end_year = None
        self.end_month = None
        self.current_parsing_year = None # To track year during parsing
        self.current_parsing_month = None # To track month during parsing

        self.analysis_results = {
            'monthly_summary': {}, # Will store opening balance, balance_after_5th etc. 
            'identified_emis': [],
            'identified_salary': [],
            # Add more keys as needed
        }

        try:
            # Assuming format MM/YYYY from your HTML datepicker 
            start_dt = datetime.strptime(start_date_str, '%m/%Y') 
            end_dt = datetime.strptime(end_date_str, '%m/%Y')

            self.start_month = start_dt.month
            self.start_year = start_dt.year
            self.end_month = end_dt.month
            self.end_year = end_dt.year

            # Initial guess for the parsing year/month 
            self.current_parsing_year = self.start_year 
            self.current_parsing_month = self.start_month

            if self.debug:
                 logger.info(f"Statement period set: Start={self.start_month}/{self.start_year}, End={self.end_month}/{self.end_year}")

        except ValueError:
             logger.error(f"Invalid date format received. Expected MM/YYYY, got start='{start_date_str}', end='{end_date_str}'") 
             # Decide how to handle: raise error, use defaults, or proceed without date logic
             raise ValueError("Invalid start or end date format provided. Please use MM/YYYY.")
        # --- End Date Processing ---

    def analyze_recurring_debits(self, min_occurrences=3, amount_tolerance=0.05, desc_similarity_threshold=80):
        """
        Analyzes transactions to identify potential recurring debits (EMIs/Loans). 
        Stores results in self.analysis_results['identified_emis']. 
        """
        logger.info("Analyzing recurring debits (potential EMIs/Loans)...") 
        self.analysis_results['identified_emis'] = []
        debits = [t for t in self.transactions if t.get('amount', 0) < 0]

        if len(debits) < min_occurrences:
            logger.info("Not enough debit transactions to analyze for recurrence.")
            return

        # Group by approximate description 
        description_groups = defaultdict(list) 
        processed_indices = set()

        for i in range(len(debits)):
            if i in processed_indices: continue
            current_desc = debits[i].get('description', '').strip().lower()
            # Remove common noise like dates or transaction IDs from description for matching 
            current_desc_cleaned = re.sub(r'\b\d{2,}/\d{2,}/\d{2,}\b|\b\d{6,}\b|[x*]{4,}', '', current_desc).strip() 
            if not current_desc_cleaned: continue

            group_key_desc = current_desc_cleaned # Use the first one as the key for the group
            description_groups[group_key_desc].append(debits[i])
            processed_indices.add(i)

            # Find similar descriptions 
            for j in range(i + 1, len(debits)):
                if j in processed_indices: continue
                compare_desc = debits[j].get('description', '').strip().lower()
                compare_desc_cleaned = re.sub(r'\b\d{2,}/\d{2,}/\d{2,}\b|\b\d{6,}\b|[x*]{4,}', '', compare_desc).strip()
                if not compare_desc_cleaned: continue

                # Use fuzzy matching 
                similarity = fuzz.token_sort_ratio(current_desc_cleaned, compare_desc_cleaned) 
                if similarity >= desc_similarity_threshold:
                    description_groups[group_key_desc].append(debits[j])
                    processed_indices.add(j)

        # Analyze each group for recurrence 
        for desc, group_transactions in description_groups.items():
            if len(group_transactions) < min_occurrences:
                continue

            # Find the most common amount(s) in the group 
            amounts = [abs(t['amount']) for t in group_transactions if t.get('amount') is not None]
            if not amounts:
                logger.warning(f"No valid amounts found in group {desc} to calculate common amount.")
                continue
            amount_counts = Counter(amounts)
            most_common_list = amount_counts.most_common(1)
            # <<< START FIX for potential unpacking issue >>>
            # most_common_amount, count = amount_counts.most_common(1)[0] # Old line
            # print(f"DEBUG ANALYSIS PRINT: ... most_common(1) result: {most_common_list}") # Your debug print 

            most_common_amount = None 
            count = 0
            unpacking_successful = False

            if most_common_list: # Check if list is not empty 
                if isinstance(most_common_list[0], (list, tuple)) and len(most_common_list[0]) == 2: # Check if first item is a pair 
                    try: 
                        most_common_amount, count = most_common_list[0] # Try unpacking
                        unpacking_successful = True
                        if self.debug: print(f"DEBUG ANALYSIS PRINT: ... Unpacking SUCCESSFUL. Amount={most_common_amount}, Count={count}") 
                    except Exception as unpack_err:
                        if self.debug: print(f"DEBUG ANALYSIS PRINT: ... UNEXPECTED error during unpacking '{most_common_list[0]}'. Error: {unpack_err}")
                        logger.error(f"Unexpected error unpacking most common amount: {unpack_err}") 
                else:
                    if self.debug: print(f"DEBUG ANALYSIS PRINT: ... Incorrect structure for unpacking: '{most_common_list[0]}'") 
                    logger.warning(f"Incorrect structure from most_common: {most_common_list[0]}")
            else:
                if self.debug: print("DEBUG ANALYSIS PRINT: ... most_common(1) returned empty list.")
                logger.warning("Could not find most common amount (empty list).") 
            # <<< END FIX >>>

            # Check if transactions with amounts close to the most common one occur regularly 
            potential_emi_group = []
            # Add check if most_common_amount was found
            if most_common_amount is not None:
                for t in group_transactions:
                    if t.get('amount') is not None and abs(abs(t['amount']) - most_common_amount) <= most_common_amount * amount_tolerance: 
                        potential_emi_group.append(t)

            if len(potential_emi_group) >= min_occurrences:
                 # Basic check for monthly recurrence (more sophisticated checks needed for robustness) 
                 potential_emi_group.sort(key=lambda x: datetime.strptime(x['date'], '%Y-%m-%d') if isinstance(x['date'], str) else x['date'])
                 month_years = set([(datetime.strptime(t['date'], '%Y-%m-%d') if isinstance(t['date'], str) else t['date']).strftime('%Y-%m') for t in potential_emi_group]) 

                 # If it occurs in multiple distinct months and frequency is roughly monthly 
                 # This is a basic heuristic - real recurrence detection is complex 
                 if len(month_years) >= min_occurrences -1 : # Allow for slight variation 
                    emi_info = {
                         'description_pattern': desc, 
                         'estimated_amount': most_common_amount,
                         'occurrences': len(potential_emi_group), 
                         'first_occurrence_date': potential_emi_group[0]['date'],
                         'last_occurrence_date': potential_emi_group[-1]['date'],
                         'transaction_ids': [t.get('transaction_id') for t in potential_emi_group] # Assuming you have IDs 
                    }
                    self.analysis_results['identified_emis'].append(emi_info) 
                    logger.info(f"Identified potential recurring debit: Desc='{desc}', Amt={most_common_amount:.2f}, Occurrences={len(potential_emi_group)}")

        if self.debug:
             logger.info(f"Identified EMIs/Loans: {self.analysis_results['identified_emis']}")

    def analyze_salary_credits(self, min_occurrences=2, amount_variance_threshold=0.15):
            """
            Analyzes transactions to identify potential salary credits. 
            Stores results in self.analysis_results['identified_salary']. 
            """
            logger.info("Analyzing potential salary credits...") 
            self.analysis_results['identified_salary'] = []
            # Consider credits that are likely significant (e.g., > certain threshold, or among the largest credits) 
            credits = sorted([t for t in self.transactions if t.get('amount', 0) > 0], key=lambda x: x['amount'], reverse=True)

            # Heuristic: Look at top N largest credits or credits above a dynamic threshold 
            potential_salary_credits = credits[:max(10, len(credits)//5)] # Look at top 10 or top 20%

            if len(potential_salary_credits) < min_occurrences:
                logger.info("Not enough significant credit transactions to analyze for salary.")
                return

            # Group by description similarity (less precise for salary maybe, but helps) 
            salary_groups = defaultdict(list)
            processed_indices_salary = set()

            # Simplified grouping: Check for keywords first 
            salary_keywords = ['salary', 'sal credit', 'wages', 'payroll', 'remuneration', 'stipend'] # Add company names if known
            keyword_based_group = []
            other_large_credits = [] 

            for i, t in enumerate(potential_salary_credits):
                desc = t.get('description','').lower()
                if any(keyword in desc for keyword in salary_keywords):
                    keyword_based_group.append(t)
                else: 
                    other_large_credits.append(t) # Keep track of other large credits

            # Analyze keyword group first 
            if len(keyword_based_group) >= min_occurrences:
                keyword_based_group.sort(key=lambda x: datetime.strptime(x['date'], '%Y-%m-%d') if isinstance(x['date'], str) else x['date'])
                # Check for regularity (e.g., mostly monthly) 
                month_years = set([(datetime.strptime(t['date'], '%Y-%m-%d') if isinstance(t['date'], str) else t['date']).strftime('%Y-%m') for t in keyword_based_group]) 
                if len(month_years) >= min_occurrences -1:
                    avg_amount = sum(t['amount'] for t in keyword_based_group) / len(keyword_based_group)
                    salary_info = { 
                        'type': 'Keyword-Based',
                        'description_sample': keyword_based_group[0].get('description'),
                        'estimated_amount': avg_amount,
                         'occurrences': len(keyword_based_group), 
                        'first_occurrence_date': keyword_based_group[0]['date'],
                        'last_occurrence_date': keyword_based_group[-1]['date'],
                        'transaction_ids': [t.get('transaction_id') for t in keyword_based_group]
                    } 
                    self.analysis_results['identified_salary'].append(salary_info)
                    logger.info(f"Identified potential salary (keyword): Avg Amt={avg_amount:.2f}, Occurrences={len(keyword_based_group)}")


            # Analyze other large credits if no keyword match or to find secondary income 
            # (More complex: needs clustering by amount/day of month) 
            # Basic version: Look for the most frequent large credit amount 
            if other_large_credits:
                amounts = [t['amount'] for t in other_large_credits]
                if amounts:
                    amount_counts = Counter(amounts)
                    most_common_list = amount_counts.most_common(1) 
                    # <<< START FIX for potential unpacking issue >>>
                    # most_common_amount, count = amount_counts.most_common(1)[0] # Old line
                    # print(f"DEBUG ANALYSIS PRINT: ... most_common(1) result: {most_common_list}") # Your debug print 

                    most_common_amount = None 
                    count = 0 
                    unpacking_successful = False

                    if most_common_list: # Check if list is not empty 
                        if isinstance(most_common_list[0], (list, tuple)) and len(most_common_list[0]) == 2: # Check if first item is a pair 
                            try: 
                                most_common_amount, count = most_common_list[0] # Try unpacking
                                unpacking_successful = True 
                                if self.debug: print(f"DEBUG ANALYSIS PRINT: ... Unpacking SUCCESSFUL. Amount={most_common_amount}, Count={count}") 
                            except Exception as unpack_err:
                                if self.debug: print(f"DEBUG ANALYSIS PRINT: ... UNEXPECTED error during unpacking '{most_common_list[0]}'. Error: {unpack_err}") 
                                logger.error(f"Unexpected error unpacking most common amount: {unpack_err}") 
                        else:
                            if self.debug: print(f"DEBUG ANALYSIS PRINT: ... Incorrect structure for unpacking: '{most_common_list[0]}'") 
                            logger.warning(f"Incorrect structure from most_common: {most_common_list[0]}") 
                    else:
                        if self.debug: print("DEBUG ANALYSIS PRINT: ... most_common(1) returned empty list.")
                        logger.warning("Could not find most common amount (empty list).") 
                    # <<< END FIX >>>

                    if most_common_amount is not None and count >= min_occurrences: 
                        # Find all transactions close to this amount
                        salary_candidates = [t for t in other_large_credits if abs(t['amount'] - most_common_amount) <= most_common_amount * amount_variance_threshold]
                        if len(salary_candidates) >= min_occurrences: 
                            salary_candidates.sort(key=lambda x: datetime.strptime(x['date'], '%Y-%m-%d') if isinstance(x['date'], str) else x['date'])
                            month_years = set([(datetime.strptime(t['date'], '%Y-%m-%d') if isinstance(t['date'], str) else t['date']).strftime('%Y-%m') for t in salary_candidates])
                            if len(month_years) >= min_occurrences - 1: 
                                salary_info = {
                                    'type': 'Amount-Based',
                                    'description_sample': salary_candidates[0].get('description'), 
                                    'estimated_amount': most_common_amount,
                                    'occurrences': len(salary_candidates),
                                    'first_occurrence_date': salary_candidates[0]['date'], 
                                    'last_occurrence_date': salary_candidates[-1]['date'],
                                    'transaction_ids': [t.get('transaction_id') for t in salary_candidates] 
                                }
                                # Avoid adding duplicate if keyword based already found similar amount/dates 
                                is_duplicate = False 
                                for existing in self.analysis_results['identified_salary']:
                                    if abs(existing['estimated_amount'] - salary_info['estimated_amount']) < 1.0: # Arbitrary small diff 
                                        is_duplicate = True 
                                        break
                                if not is_duplicate: 
                                    self.analysis_results['identified_salary'].append(salary_info)
                                    logger.info(f"Identified potential salary (amount): Amt={most_common_amount:.2f}, Occurrences={len(salary_candidates)}")

            if self.debug: 
                logger.info(f"Identified Salary: {self.analysis_results['identified_salary']}")

    def perform_analysis(self):
            """Runs all analytical functions."""
            logger.info("DEBUG STEP 5: Inside perform_analysis (PARTIAL)") 
            if not self.transactions:
                logger.warning("No transactions found, skipping analysis.")
                return 

            # Ensure balances are calculated first if needed by other analyses 
            # Assuming _ensure_running_balance() was called previously if needed 
            logger.info("DEBUG STEP 5: Calling _calculate_monthly_balances() from perform_analysis") 
            self._calculate_monthly_balances()

            # Run specific analyses 
            # logger.info("DEBUG STEP 5: SKIPPING analyze_recurring_debits")  # Commented out skip
            self.analyze_recurring_debits() # Find EMIs/Loans 
            logger.info("DEBUG STEP 5: Calling analyze_salary_credits()") # Updated log 
            self.analyze_salary_credits()   # Find Salary

            logger.info("DEBUG STEP 5: All analysis complete.") 
            # The results are stored in self.analysis_results


    def extract_all(self):
        """Extracts all information and performs analysis."""
        # 1. Extract raw text for metadata 
        logger.info("DEBUG STEP 3: Calling extract_text()") # Add debug print 
        raw_text = self.extract_text()

        # 2. Extract account information from raw text 
        logger.info("DEBUG STEP 3: Calling extract_account_info()") # Add debug print 
        self.extract_account_info(raw_text)

        # 3. Extract transactions 
        # logger.info("DEBUG STEP 3: SKIPPING extract_transactions and perform_analysis")  # Commented out skip
        self.extract_transactions() # Keep original function 
        # 4. Perform Analysis
        self.perform_analysis() # Call analysis after transactions are extracted

        logger.info(f"EXTRACTOR DEBUG: Before returning from extract_all - self.transactions length: {len(self.transactions)}") 
        return list(self.transactions), self.account_info, self.analysis_results 

    def extract_text(self):
        """Extract raw text from PDF for metadata processing, handling password if provided.""" 
        text = ""
        try:
            logger.info(f"Extracting text from {self.pdf_path}") 
            with open(self.pdf_path, 'rb') as file: 
                reader = PyPDF2.PdfReader(file)

                # Handle encrypted PDF 
                if reader.is_encrypted:
                    # Try empty string password first as some PDFs might use it 
                    if self.password is None:
                        logger.warning("PDF is encrypted, but no password provided. Trying empty password.")
                        if not reader.decrypt(''):
                            logger.error("Decryption with empty password failed. Password required.")
                            raise ValueError("PDF is encrypted and requires a password")
                    elif not reader.decrypt(self.password):
                        logger.error("Incorrect password provided for encrypted PDF.")
                        raise ValueError("Incorrect password for encrypted PDF") 

                    # After decryption attempt, try to access pages and extract text
                    page_accessible = False
                    try:
                        # Try to access and extract text from at least one page
                        for page_num in range(len(reader.pages)):
                            try:
                                _ = reader.pages[page_num]
                                # Try extracting text as well
                                _ = reader.pages[page_num].extract_text()
                                page_accessible = True
                                break
                            except Exception:
                                continue
                        if page_accessible:
                            logger.info("PDF successfully decrypted.")
                        else:
                            logger.error("PDF decryption failed after password attempt. Possibly incorrect password or corrupted PDF.")
                            raise ValueError("PDF decryption failed. Check password or PDF integrity.")
                    except Exception as e:
                        logger.error("PDF decryption failed after password attempt. Possibly incorrect password or corrupted PDF.")
                        raise ValueError("PDF decryption failed. Check password or PDF integrity.")

                # Check if decryption worked (needed even if no password was provided initially) 
                # if reader.is_encrypted:
                #     # This check is important if decryption failed silently 
                #     logger.error("PDF decryption failed, possibly due to incorrect logic or library issue.")
                #     raise ValueError("PDF decryption failed. Check password or PDF integrity.")

                for page_num, page in enumerate(reader.pages): 
                    try:
                        page_text = page.extract_text()
                        if page_text: # Ensure text was actually extracted 
                            text += page_text + "\n"
                            if self.debug:
                                logger.info(f"Page {page_num+1} extracted {len(page_text)} characters") 
                        else:
                            if self.debug:
                                logger.warning(f"Page {page_num+1} extracted no text (possibly image-based page or extraction error)") 
                    except Exception as page_error:
                        logger.error(f"Error extracting text from page {page_num+1}: {str(page_error)}") 
                        # Optionally continue to next page or re-raise

        except FileNotFoundError as fnf:
              logger.error(f"PDF file not found at path: {self.pdf_path}") 
              raise fnf # Re-raise the specific error
        except ValueError as ve: # Specific errors like password issues 
             logger.error(f"PDF processing error: {ve}")
             raise ve # Re-raise
        except Exception as e:
            logger.error(f"General error extracting text from PDF: {str(e)}", exc_info=self.debug) # Show traceback if debug 
            raise Exception(f"Error extracting text from PDF: {str(e)}") 

        if not text and self.debug:
             logger.warning("No text extracted from the entire PDF. It might be purely image-based or corrupted.") 

        if self.debug: 
            # Save the extracted text to file for debugging
            try:
                debug_text_path = f"{os.path.splitext(self.pdf_path)[0]}_{self.unique_id}_extracted_text.txt"
                with open(debug_text_path, 'w', encoding='utf-8') as f:
                    f.write(text) 
                logger.info(f"Saved extracted text to {debug_text_path} for debugging")
            except Exception as write_error:
                logger.error(f"Failed to write debug text file: {write_error}")

        return text

    # ============================================================
    # ========= UPDATED extract_account_info Method Start ========
    # ============================================================
    def extract_account_info(self, text):
        """Extract account information from the bank statement text using regex."""
        logger.info("Extracting account information") 
        self.account_info = {} # Reset account info for this extraction

        # Priority 1: User-provided bank name
        # (Already handled in __init__)

        # Use a set to store potential candidates and avoid duplicates
        account_number_candidates = set()

        # --- Priority 2: Regex patterns targeting specific user constraints ---
        logger.debug("Account Info: Applying priority regex patterns (numeric, length, masking)")
        # Pattern: Keyword followed by 8-19 digits (common format, less than 20 digits)
        # Allows for optional spaces or hyphens within the number
        pattern_strict_numeric = re.compile(
            r'(?:Account\s*(?:Number|#|No\.?)|A/C\s*No\.?)\s*[:\-]?\s*(\d(?:[\s\-]?\d){7,18})\b',
            re.IGNORECASE
        )
        for match in pattern_strict_numeric.finditer(text):
            num_str = match.group(1)
            cleaned_num = re.sub(r'[\s\-]', '', num_str) # Remove spaces/hyphens
            if cleaned_num.isdigit() and 8 <= len(cleaned_num) <= 19:
                logger.info(f"Found account number candidate (Strict Numeric): {cleaned_num}")
                account_number_candidates.add(cleaned_num)

        # Pattern: Masked numbers (e.g., XXXXXX1234, ******1234) - allows 4+ mask chars followed by 3+ digits
        pattern_masked = re.compile(r'\b([Xx\*]{4,}\d{3,})\b')
        for match in pattern_masked.finditer(text):
            masked_num = match.group(1)
            logger.info(f"Found account number candidate (Masked): {masked_num}")
            account_number_candidates.add(masked_num) # Store masked numbers as well

        # --- Priority 3: Existing broader regex patterns (if no match above) ---
        if not account_number_candidates:
            logger.debug("Account Info: Priority patterns failed, trying broader regex patterns")
            existing_account_patterns = [
                # Standard formats with various labels 
                r'(?:Account|A/C|A/c|Acc\.?)\s*(?:Number|#|No\.?|No|Num)\s*:?\s*([A-Z0-9\-\s\.]+)',
                # Look for standalone account numbers 
                r'\b(?:Account|A/C|A/c|Acc\.?)\s*(?:Number|#|No\.?|No|Num)\b[^a-zA-Z0-9]*([A-Z0-9\-\s\.]{8,})',
                # Common account number formats without labels 
                r'\b(\d{10,18})\b',  # 10-18 digit numbers
                r'\b(\d{4}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4})\b',  # 4-4-4-4 format 
                r'\b(\d{3}[-\s]?\d{7}[-\s]?\d{2})\b',  # 3-7-2 format 
                # r'\b[A-Z0-9]{8,12}\b',  # 8-12 alphanumeric characters (Can be too broad, keep commented unless needed) 
            ]
            for pattern in existing_account_patterns:
                try:
                    matches = re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE)
                    for match in matches:
                        value = match.group(1).strip().replace('\n', ' ')
                        # Basic validation (skip dates, amounts, very short strings)
                        if re.match(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', value) or \
                           re.match(r'^\(?[\d,.-]+\)?$', value) or \
                           len(re.sub(r'[-\s]', '', value)) < 8:
                            continue
                        # Further check for pure numeric < 20 digits (preference)
                        cleaned_val = re.sub(r'[-\s]', '', value)
                        if cleaned_val.isdigit() and len(cleaned_val) <= 19:
                             logger.info(f"Found account number candidate (Existing Numeric): {cleaned_val}")
                             account_number_candidates.add(cleaned_val)
                        elif 'X' in value or '*' in value: # Accept masked from existing patterns too
                             logger.info(f"Found account number candidate (Existing Masked): {value}")
                             account_number_candidates.add(value)
                        # Optional: Add alphanumeric if needed, but with caution
                        # elif re.match(r'^[A-Z0-9]+$', cleaned_val) and len(cleaned_val) <= 19:
                        #    logger.info(f"Found account number candidate (Existing AlphaNum): {cleaned_val}")
                        #    account_number_candidates.add(cleaned_val)

                except re.error as re_err:
                    logger.warning(f"Regex error for account number pattern '{pattern}': {re_err}")

        # --- Selection Logic ---
        selected_account_number = None
        if account_number_candidates:
            logger.debug(f"Account number candidates: {account_number_candidates}")
            # Prioritize purely numeric candidates within the length limit
            numeric_candidates = {num for num in account_number_candidates if num.isdigit() and len(num) <= 19}
            # Prioritize longer numeric candidates if multiple exist
            if numeric_candidates:
                selected_account_number = max(numeric_candidates, key=len)
                logger.info(f"Selected final account number (Numeric): {selected_account_number}")
            else:
                # If no purely numeric found, take any candidate (might be masked)
                # Prioritize longer candidates among the remaining ones
                selected_account_number = max(account_number_candidates, key=lambda x: len(re.sub(r'[-\s]', '', x)))
                logger.info(f"Selected final account number (Fallback/Masked): {selected_account_number}")
        else:
            logger.warning("Could not extract account number using any regex method.")

        self.account_info['account_number'] = selected_account_number

        # --- Extract Other Information (Keep existing logic) ---
        logger.debug("Account Info: Extracting other details (name, period, balances etc.)")
        other_patterns = {
            # Statement Info
            'statement_period': r'Statement\s?Period(?:\s*From)?:\s*([A-Za-z0-9\s,./\-]+(?:to|-)\s*[A-Za-z0-9\s,./\-]+)', 
            'statement_date': r'Statement\s?Date:\s*([A-Za-z0-9\s,./\-]+)', 
            # Balances
            'opening_balance': r'(?:Opening|Previous|Beginning)\s?Balance:?\s*[$€£]?\s*([\d,.-]+)', 
            'closing_balance': r'(?:Closing|Ending)\s?Balance:?\s*[$€£]?\s*([\d,.-]+)', 
            'available_balance': r'Available\s?Balance:?\s*[$€£]?\s*[$€£]?\s*([\d,.-]+)', 
            # Account/Owner Details
            'account_name': r'Account\s?(?:Name|Holder):\s*([A-Za-z0-9\s\.\-\'&]+)', 
            'owner_name': r'^(?:Mr\.?|Mrs\.?|Ms\.?|Dr\.?)\s*([A-Za-z\s.\-\']+)\s*$', 
            'customer_name': r'(?:Customer|Prepared\s+for):\s*([A-Za-z0-9\s\.\-\'&]+)', 
            'account_type': r'Account\s?Type:?\s*([A-Za-z\s]+(?:Account)?)', 
            # Bank/Branch Details
            'bank_name': r'(KOTAK\s+MAHINDRA\s+BANK|HDFC\s+BANK|ICICI\s+BANK|STATE\s+BANK\s+OF\s+INDIA|SBI|AXIS\s+BANK|YES\s+BANK|IDFC\s+BANK|BANDHAN\s+BANK|RBL\s+BANK|FEDERAL\s+BANK|SOUTH\s+INDIAN\s+BANK|KARNATAKA\s+BANK|CANARA\s+BANK|PUNJAB\s+NATIONAL\s+BANK|BANK\s+OF\s+BARODA|UNION\s+BANK|BANK\s+OF\s+INDIA|CENTRAL\s+BANK|INDIAN\s+BANK|UCO\s+BANK|PUNJAB\s+& SIND\s+BANK|BANK\s+OF\s+MAHARASHTRA|ANDHRA\s+BANK|VIJAYA\s+BANK|CORPORATION\s+BANK|SYNDICATE\s+BANK|ORIENTAL\s+BANK|UNITED\s+BANK|ALLAHABAD\s+BANK|DENA\s+BANK)',
            'branch_code': r'Branch\s?(?:Code|No\.?):\s*([A-Z0-9-]+)', 
            'routing_number': r'(?:Routing|ABA)\s?(?:Number|#|No\.?):\s*([0-9]{9})', 
            'swift_code': r'SWIFT\s?(?:Code|BIC):\s*([A-Z0-9]{8,11})', 
            'iban': r'IBAN:\s*([A-Z0-9\s]+)', 
            # Other Fields
            'customer_id': r'Customer\s*ID:?\s*([A-Z0-9\-]+)', 
            'tax_id': r'(?:Tax\s*ID|SSN|EIN):?\s*([A-Z0-9\-]+)', 
            'credit_limit': r'Credit\s?Limit:?\s*[$€£]?\s*([\d,.]+)', 
            'interest_rate': r'Interest\s?Rate:?\s*([\d.]+)%?', 
            'minimum_balance': r'Minimum\s?Balance:?\s*[$€£]?\s*([\d,.]+)', 
        }

        for key, pattern in other_patterns.items():
             # Only search if not already found (avoid overwriting good matches)
             # Special handling for bank_name: don't override user's selection
             if key not in self.account_info or (key == 'bank_name' and not self.account_info.get('bank_name')):
                try:
                    match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
                    if match:
                        value = match.group(1).strip().replace('\n', ' ')
                        if 'balance' in key: 
                            value = self._clean_balance_string(value) 
                        
                        # CRITICAL FIX: Never override user's bank name selection
                        if key == 'bank_name' and self.account_info.get('bank_name'):
                            if self.debug:
                                logger.info(f"Skipping bank_name extraction - user already selected: {self.account_info['bank_name']}")
                            continue
                        
                        self.account_info[key] = value
                        if self.debug:
                            logger.info(f"Found {key}: {value}") 
                except re.error as re_err:
                    logger.warning(f"Regex error for key '{key}': {re_err}") 


        # --- Fallbacks and Inferences (Keep existing logic) --- 
        # If owner_name not found, use account_name or customer_name 
        if 'owner_name' not in self.account_info:
            if 'account_name' in self.account_info:
                 self.account_info['owner_name'] = self.account_info['account_name']
            elif 'customer_name' in self.account_info:
                 self.account_info['owner_name'] = self.account_info['customer_name']

        # Infer account type if not found 
        if 'account_type' not in self.account_info:
            account_type_keywords = {
                'Checking': ['checking', 'cheque', 'current account'],
                'Savings': ['saving', 'deposit account'],
                'Credit Card': ['credit card', 'visa', 'mastercard', 'amex'],
                'Loan': ['loan', 'mortgage', 'lending'], 
                'Investment': ['investment', 'brokerage'],
            }
            # Search in the first few lines or near account number for clues 
            lines = text.split('\n')
            search_text_type = "\n".join(lines[:15])
            for acc_type, keywords in account_type_keywords.items(): 
                if any(re.search(r'\b' + keyword + r'\b', search_text_type, re.IGNORECASE) for keyword in keywords):
                    self.account_info['account_type'] = acc_type
                    logger.info(f"Inferred account type: {acc_type}")
                    break # Stop after first match

        # Store extraction ID
        self.account_info['extraction_id'] = self.unique_id 

        # Fallback bank name extraction if not found by regex
        # Only auto-detect if user hasn't provided a bank name
        if 'bank_name' not in self.account_info or not self.account_info['bank_name']:
            extracted_bank = self._extract_bank_name_fallback(text)
            if extracted_bank:
                self.account_info['bank_name'] = extracted_bank
                if self.debug:
                    logger.info(f"Extracted bank name via fallback: {extracted_bank}")
        else:
            if self.debug:
                logger.info(f"Using user-provided bank name: {self.account_info['bank_name']}")

        # Log missing essential fields 
        essential_fields = ['account_number', 'owner_name', 'bank_name', 'statement_period'] 
        for field in essential_fields:
            if not self.account_info.get(field) and self.debug: # Use .get() for safety
                logger.warning(f"Essential field '{field}' could not be extracted.")
    # ============================================================
    # ========= UPDATED extract_account_info Method End ==========
    # ============================================================

    def _extract_bank_name_fallback(self, text: str) -> Optional[str]:
        """Fallback method to extract bank name from text content."""
        text_lower = text.lower()
        
        # Common bank name patterns
        bank_patterns = {
            'KOTAK MAHINDRA BANK': ['kotak mahindra bank', 'kmbl', 'kotak.com', 'kotak bank'],
            'HDFC BANK': ['hdfc bank', 'hdfc.com', 'hdfc'],
            'ICICI BANK': ['icici bank', 'icici.com', 'icici'],
            'STATE BANK OF INDIA': ['state bank of india', 'sbi', 'sbi.co.in'],
            'AXIS BANK': ['axis bank', 'axis.com', 'axis'],
            'YES BANK': ['yes bank', 'yesbank.in', 'yes'],
            'IDFC BANK': ['idfc bank', 'idfc.com', 'idfc'],
            'BANDHAN BANK': ['bandhan bank', 'bandhanbank.com', 'bandhan'],
            'RBL BANK': ['rbl bank', 'rblbank.com', 'rbl'],
            'FEDERAL BANK': ['federal bank', 'federalbank.co.in', 'federal'],
            'SOUTH INDIAN BANK': ['south indian bank', 'southindianbank.com', 'south indian'],
            'KARNATAKA BANK': ['karnataka bank', 'karnatakabank.com', 'karnataka'],
            'CANARA BANK': ['canara bank', 'canarabank.com', 'canara'],
            'PUNJAB NATIONAL BANK': ['punjab national bank', 'pnb.co.in', 'pnb'],
            'BANK OF BARODA': ['bank of baroda', 'bankofbaroda.com', 'bob'],
            'UNION BANK': ['union bank', 'unionbankofindia.co.in', 'union'],
            'BANK OF INDIA': ['bank of india', 'bankofindia.com', 'boi'],
            'CENTRAL BANK': ['central bank', 'centralbank.co.in', 'central'],
            'INDIAN BANK': ['indian bank', 'indianbank.co.in', 'indian'],
            'UCO BANK': ['uco bank', 'ucobank.com', 'uco'],
            'PUNJAB & SIND BANK': ['punjab & sind bank', 'psbindia.com', 'psb'],
            'BANK OF MAHARASHTRA': ['bank of maharashtra', 'bankofmaharashtra.in', 'bom'],
            'ANDHRA BANK': ['andhra bank', 'andhrabank.in', 'andhra'],
            'VIJAYA BANK': ['vijaya bank', 'vijayabank.com', 'vijaya'],
            'CORPORATION BANK': ['corporation bank', 'corporationbank.com', 'corporation'],
            'SYNDICATE BANK': ['syndicate bank', 'syndicatebank.in', 'syndicate'],
            'ORIENTAL BANK': ['oriental bank', 'obcindia.co.in', 'obc'],
            'UNITED BANK': ['united bank', 'unitedbankofindia.com', 'united'],
            'ALLAHABAD BANK': ['allahabad bank', 'allahabadbank.in', 'allahabad'],
            'DENA BANK': ['dena bank', 'denabank.com', 'dena']
        }
        
        # Search for bank names in text
        for bank_name, patterns in bank_patterns.items():
            for pattern in patterns:
                if pattern in text_lower:
                    return bank_name
        
        # If no exact match, try to find partial matches
        for bank_name, patterns in bank_patterns.items():
            for pattern in patterns:
                if any(word in text_lower for word in pattern.split()):
                    return bank_name
        
        return None

    def _clean_balance_string(self, balance_str):
        """Removes currency symbols, thousand separators, and handles negative formats.""" 
        if not balance_str: return ''
        cleaned = re.sub(r'[$\s€£,]', '', balance_str) # Remove common symbols and spaces 
        # Handle trailing minus or CR/DR indicators if needed 
        if cleaned.endswith('CR'):
            cleaned = cleaned[:-2]
        elif cleaned.endswith('DR'):
             cleaned = '-' + cleaned[:-2]
        elif cleaned.endswith('-'):
             cleaned = '-' + cleaned[:-1]
        # Handle parentheses for negatives 
        if cleaned.startswith('(') and cleaned.endswith(')'):
             cleaned = '-' + cleaned[1:-1]
        return cleaned

    def extract_transactions(self):
        tabula_options = {
            "pages": "all",
            "multiple_tables": True,
            "guess": True,
            "lattice": True,
            "stream": True,
            "password": self.password
        }
        first_page_text = self._extract_first_page_text(self.pdf_path)
        template = self._detect_template(first_page_text)
        tables = read_pdf(self.pdf_path, **tabula_options)
        parser_func = getattr(self, TEMPLATE_PARSERS.get(template, '_parse_generic'))
        all_transactions = []
        for df in tables:
            all_transactions.extend(parser_func(df))
        self.transactions = all_transactions
        # Populate high value transactions
        self.high_value_transactions = [
            t for t in self.transactions
            if t.get('amount') is not None and abs(t.get('amount', 0)) >= self.high_value_threshold
        ]

    def _extract_first_page_text(self, pdf_path):
        try:
            with pdfplumber.open(pdf_path) as pdf:
                return pdf.pages[0].extract_text()
        except Exception:
            return ''

    def _detect_template(self, first_page_text):
        text = first_page_text.lower() if first_page_text else ''
        if "withdrawal(dr)/deposit(cr)" in text:
            return "KOTAK_SINGLE_COL"
        if "debit" in text and "credit" in text and "transaction details" in text:
            return "HDFC_SEPARATE_COLS"
        # Add more detection rules for other banks/templates
        return "GENERIC"

    def _parse_kotak_single_col(self, df):
        transactions = []
        for _, row in df.iterrows():
            date = row.get('Date')
            desc = row.get('Narration')
            ref = row.get('Chq/Ref No')
            amount_str = str(row.get('Withdrawal(Dr)/Deposit(Cr)', '')).replace(',', '').strip()
            balance_str = str(row.get('Balance', '')).replace(',', '').strip()
            # Parse amount
            if amount_str.endswith('(Dr)'):
                amount = -float(amount_str[:-4])
            elif amount_str.endswith('(Cr)'):
                amount = float(amount_str[:-4])
            else:
                amount = None
            # Parse balance
            if balance_str.endswith('(Dr)'):
                balance = -float(balance_str[:-4])
            elif balance_str.endswith('(Cr)'):
                balance = float(balance_str[:-4])
            else:
                balance = None
            transactions.append({
                'date': date,
                'description': desc,
                'ref': ref,
                'amount': amount,
                'balance': balance
            })
        return transactions



    def _parse_hdfc_separate_cols(self, df):
        transactions = []
        for _, row in df.iterrows():
            date = row.get('DATE')
            desc = row.get('TRANSACTION DETAILS')
            ref = row.get('CHEQUE/REFERENCE#')
            debit = row.get('DEBIT')
            credit = row.get('CREDIT')
            balance = row.get('BALANCE')
            # Parse amounts
            amount = None
            if debit and str(debit).strip():
                amount = -float(str(debit).replace(',', '').strip())
            elif credit and str(credit).strip():
                amount = float(str(credit).replace(',', '').strip())
            # Parse balance
            try:
                balance = float(str(balance).replace(',', '').strip())
            except Exception:
                balance = None
            transactions.append({
                'date': date,
                'description': desc,
                'ref': ref,
                'amount': amount,
                'balance': balance
            })
        return transactions

    def _parse_generic(self, df):
        """Robust generic parser for unknown templates. Attempts to extract transactions from any table."""
        transactions = []
        if df is None or df.empty:
            logger.info("[GENERIC PARSER] DataFrame is empty, skipping.")
            return []

        # Clean columns (handle unnamed, numeric, or bad headers)
        df = self._clean_table_columns(df)
        if df is None or df.empty:
            logger.info("[GENERIC PARSER] DataFrame could not be cleaned, skipping.")
            return []

        # Identify likely columns
        col_map = self._identify_columns(df)
        logger.info(f"[GENERIC PARSER] Detected columns: {col_map}")
        if not col_map.get('date') or not (col_map.get('amount') or (col_map.get('debit') and col_map.get('credit'))):
            logger.info("[GENERIC PARSER] Required columns not found, skipping table.")
            return []

        for idx, row in df.iterrows():
            # Parse date
            date_val = row.get(col_map.get('date'), '')
            date = self._parse_date(date_val)
            if not date:
                continue  # Skip rows without a valid date

            # Parse description
            desc = row.get(col_map.get('desc'), '') if col_map.get('desc') else ''
            if pd.isna(desc):
                desc = ''
            desc = str(desc).strip()

            # Parse amount
            amount = None
            if col_map.get('amount'):
                amount = self._parse_amount(row.get(col_map['amount'], ''))
            elif col_map.get('debit') and col_map.get('credit'):
                debit = self._parse_amount(row.get(col_map['debit'], ''))
                credit = self._parse_amount(row.get(col_map['credit'], ''))
                if debit is not None and debit != 0:
                    amount = -abs(debit)
                elif credit is not None and credit != 0:
                    amount = abs(credit)
            if amount is None:
                continue  # Skip rows without a valid amount

            # Parse balance
            balance = None
            if col_map.get('balance'):
                balance = self._parse_amount(row.get(col_map['balance'], ''))

            # Parse reference if available
            ref = row.get('ref', '') or row.get('Ref', '') or row.get('Reference', '') or row.get('Chq/Ref No', '')
            if pd.isna(ref):
                ref = ''
            ref = str(ref).strip()

            # Categorize transaction
            category = self._categorize_transaction(desc, amount)

            transactions.append({
                'date': date,
                'description': desc,
                'ref': ref,
                'amount': amount,
                'balance': balance,
                'category': category
            })
        logger.info(f"[GENERIC PARSER] Extracted {len(transactions)} transactions from table.")
        return transactions

    def _ensure_running_balance(self):
        """Calculates running balance if missing or checks consistency.""" 
        if not self.transactions: return

        # Check if ANY transaction has a balance value first
        balance_col_parsed = any(t.get('balance') is not None for t in self.transactions)

        if not balance_col_parsed:
             logger.info("No running balance column found/parsed. Attempting calculation from opening balance.") 
             consistent = False # Force recalculation
        else:
             consistent = True
             # Start check from second transaction 
             if len(self.transactions) > 1:
                try: # Add try-except for safety 
                     for i in range(1, min(len(self.transactions), 5)): # Check first few 
                        prev_balance = self.transactions[i-1].get('balance')
                        current_amount = self.transactions[i].get('amount')
                        # Ensure both are valid numbers before arithmetic
                        if prev_balance is None or current_amount is None:
                             logger.warning(f"Cannot check consistency at index {i}, missing previous balance or current amount.")
                             consistent = False # Treat as inconsistent if values are missing
                             break
                        expected_next = round(prev_balance + current_amount, 2) 
                        curr_bal = self.transactions[i].get('balance') 
                        if curr_bal is None: # If current is missing, we can't check
                            consistent = False
                            logger.warning(f"Cannot check consistency at index {i}, current balance missing.")
                            break
                        if abs(curr_bal - expected_next) > 0.05: # Allow for small float differences (increased tolerance slightly)
                            consistent = False
                            logger.warning(f"Running balance column seems inconsistent at index {i}. Prev: {prev_balance}, Amt: {current_amount}, Curr: {curr_bal}, Expected: {expected_next}") 
                            break # Stop checking on first inconsistency
                except Exception as e:
                    logger.error(f"Error checking balance consistency: {e}", exc_info=self.debug) 
                    consistent = False

             if consistent:
                 logger.info("Running balance column found and seems consistent.") 
                 return # Balances seem okay

        # Recalculate if missing or inconsistent 
        logger.info("Recalculating running balance.")
        running_balance = None
        # Try using opening balance stored previously
        # Use self.statement_opening_balance (which should be float/None)
        # or try parsing from self.account_info['opening_balance'] again
        if self.statement_opening_balance is not None:
             running_balance = self.statement_opening_balance
             logger.info(f"Starting balance calculation with stored statement opening balance: {running_balance}") 
        else:
            opening_balance_str = self._clean_balance_string(self.account_info.get('opening_balance', '')) 
            try:
                if opening_balance_str:
                    running_balance = float(opening_balance_str) 
                    logger.info(f"Starting balance calculation with Opening Balance from account_info: {running_balance}")
                else:
                    logger.warning("No opening balance available from account_info.")
            except (ValueError, TypeError):
                logger.warning("Could not parse opening balance from account_info.") 

        # Fallback: if still no starting balance, try to infer from first transaction
        if running_balance is None: 
             if self.transactions and self.transactions[0].get('balance') is not None and self.transactions[0].get('amount') is not None:
                  # Balance *before* first transaction = balance *after* - amount
                  running_balance = round(self.transactions[0]['balance'] - self.transactions[0]['amount'], 2)
                  logger.info(f"Starting balance calculation based on first transaction's balance. Inferred starting balance: {running_balance}") 
             else:
                  logger.error("Cannot determine starting point for running balance calculation. Balance calculation aborted.") 
                  return # Cannot proceed

        # Calculate forward 
        for i, transaction in enumerate(self.transactions):
             current_amount = transaction.get('amount')
             if running_balance is None:
                 # Attempt to re-sync if we lost track but find a balance later
                 current_balance = transaction.get('balance')
                 if current_balance is not None and current_amount is not None: 
                     running_balance = round(current_balance - current_amount, 2)
                     logger.warning(f"Re-synced running balance calculation at index {i}. Inferred balance before this txn: {running_balance}")
                 else:
                     # Still cannot determine balance
                     transaction['balance'] = None
                     if i < 5 and self.debug: logger.warning(f"Cannot calculate running balance for transaction index {i}, missing prior balance.") 
                     continue # Skip to next transaction if balance cannot be calculated

             # We have a running balance, add current amount
             if current_amount is not None:
                 running_balance += current_amount 
                 transaction['balance'] = round(running_balance, 2) # Update the balance 
             else:
                 # If amount is missing, cannot continue calculation accurately
                 transaction['balance'] = None
                 running_balance = None # Lose track
                 logger.warning(f"Transaction index {i} has no amount, cannot continue running balance calculation past this point.")

    def _calculate_monthly_balances(self):
        """
        Calculate account balance on/closest after the 5th, 15th, and 25th of each month
        and the opening balance for the month based on the first transaction.
        """
        self.analysis_results['monthly_summary'] = {}
        self.monthly_balances = {}

        if not self.transactions or not any(t.get('balance') is not None for t in self.transactions):
            logger.warning("Cannot calculate monthly balances: No transactions or running balance unavailable.")
            return

        logger.info("Calculating monthly opening balance and balances on/after 5th, 15th, 25th.")
        monthly_data = defaultdict(list)
        for transaction in self.transactions:
            try:
                if isinstance(transaction.get('date'), str):
                    date_obj = datetime.strptime(transaction['date'], '%Y-%m-%d')
                elif isinstance(transaction.get('date'), datetime):
                    date_obj = transaction['date']
                else:
                    logger.warning(f"Skipping transaction with unexpected date type: {type(transaction.get('date'))}")
                    continue
                month_key = f"{date_obj.year}-{date_obj.month:02d}"
                monthly_data[month_key].append({'data': transaction, 'date_obj': date_obj})
            except (ValueError, TypeError, KeyError) as e:
                logger.warning(f"Skipping transaction with invalid date/structure for monthly balance calc: {transaction.get('date')}, Error: {e}")

        for month_key, transactions_in_month in monthly_data.items():
            transactions_in_month = sorted(transactions_in_month, key=lambda x: x['date_obj'])
            year, month = map(int, month_key.split('-'))
            target_dates = {
                '5th': datetime(year, month, 5),
                '15th': datetime(year, month, 15),
                '25th': datetime(year, month, 25)
            }

            opening_balance_month = None
            opening_balance_date = None
            balances_after = {'5th': (None, None), '15th': (None, None), '25th': (None, None)}

            # Opening balance logic (as before)
            if transactions_in_month:
                first_trans_item = transactions_in_month[0]
                first_trans_data = first_trans_item['data']
                desc = first_trans_data.get('description', '').lower()
                if any(k in desc for k in ['opening balance', 'bal b/f', 'brought forward', 'balance forward']):
                    opening_balance_month = first_trans_data.get('balance')
                    opening_balance_date = first_trans_data['date']
                    first_trans_data['category'] = 'Opening Balance'
                elif first_trans_data.get('balance') is not None and first_trans_data.get('amount') is not None:
                    opening_balance_month = round(first_trans_data['balance'] - first_trans_data['amount'], 2)
                    opening_balance_date = first_trans_data['date']

            # For each target date (5th, 15th, 25th), find the balance on or after that date
            for label, target_date in target_dates.items():
                found = False
                for item in transactions_in_month:
                    trans_data = item['data']
                    if item['date_obj'] >= target_date and trans_data.get('balance') is not None:
                        balances_after[label] = (trans_data['balance'], trans_data['date'])
                        found = True
                        break
                if not found:
                    # Fallback: use last transaction before the target date, or opening balance
                    transactions_before = [item for item in transactions_in_month if item['date_obj'] < target_date and item['data'].get('balance') is not None]
                    if transactions_before:
                        last_before = transactions_before[-1]['data']
                        balances_after[label] = (last_before['balance'], last_before['date'])
                    else:
                        balances_after[label] = (opening_balance_month, opening_balance_date)

            self.analysis_results['monthly_summary'][month_key] = {
                'opening_balance': opening_balance_month,
                'opening_balance_date': opening_balance_date,
                'balance_after_5th': balances_after['5th'][0],
                'balance_after_5th_date': balances_after['5th'][1],
                'balance_after_15th': balances_after['15th'][0],
                'balance_after_15th_date': balances_after['15th'][1],
                'balance_after_25th': balances_after['25th'][0],
                'balance_after_25th_date': balances_after['25th'][1]
            }

    def _is_transaction_table(self, df):
        """Heuristic check if a DataFrame looks like a transaction table.""" 
        if df.empty: return False

        # Lowercase column names for easier matching 
        cols = [str(col).lower().strip() for col in df.columns]

        # Check for common header terms 
        header_terms = ['date', 'desc', 'narrat', 'particular', 'detail', 'amount', 'debit', 'credit', 'payment', 'deposit', 'withdraw', 'balance'] 
        if sum(any(term in col for term in header_terms) for col in cols) >= 2: # Need at least 2 common terms 
            return True

        # Check content if headers are ambiguous (e.g., 'Unnamed: 0') 
        # Look for date pattern in first few columns 
        date_pattern = r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}\s+[A-Za-z]{3}'
        amount_pattern = r'[$€£]?[\d,.-]+' # Matches numbers with commas, dots, currency symbols 

        date_col_found = False 
        amount_col_found = False

        for i in range(min(3, len(df.columns))): # Check first 3 columns for dates 
             # Use .get() for safe access in case column index doesn't exist
             col_data = df.iloc[:, i].astype(str).dropna().head(5) # Sample first 5 non-NA values 
             if col_data.empty: continue
             if sum(col_data.str.contains(date_pattern, regex=True)) >= 2: # At least 2 date-like entries 
                  date_col_found = True 
                  break

        for i in range(len(df.columns)): # Check all columns for amounts 
             col_data = df.iloc[:, i].astype(str).dropna().head(5)
             if col_data.empty: continue
             if sum(col_data.str.contains(amount_pattern, regex=True)) >= 2: # At least 2 amount-like entries 
                  amount_col_found = True
                  break

        return date_col_found and amount_col_found

    def _clean_table_columns(self, df):
        """Cleans DataFrame columns, potentially using the first row as headers.""" 
        clean_df = df.copy()

        # --- Stage 1: Handle completely numeric or unnamed columns --- 
        # If all columns are ints or 'Unnamed: X', the first row is likely the header 
        if all(isinstance(col, int) or str(col).startswith('Unnamed:') for col in clean_df.columns):
            if not clean_df.empty:
                first_row_values = [str(val).strip() for val in clean_df.iloc[0]] 
                # Basic check: Do these look like headers? (Not just numbers or empty) 
                if any(re.search(r'[a-zA-Z]', h) for h in first_row_values if h): # Contains letters 
                    logger.info("Using first row as header because columns were unnamed/numeric.")
                    clean_df.columns = first_row_values
                    clean_df = clean_df.iloc[1:].reset_index(drop=True) 
                else:
                    # Headers are still bad, maybe assign generic names? 
                    logger.warning("Columns are unnamed/numeric, but first row doesn't look like headers. Assigning generic names.") 
                    clean_df.columns = [f'col_{i}' for i in range(len(clean_df.columns))]
            else:
                 logger.warning("Cannot clean columns of empty unnamed DataFrame.") 
                 return None # Cannot process further

        # --- Stage 2: Clean up existing string column names --- 
        new_columns = []
        for col in clean_df.columns:
            if pd.isna(col): # Handle potential NaN column names 
                new_name = f"unnamed_{len(new_columns)}"
            else:
                new_name = str(col).strip().replace('\n', ' ').replace('\r', ' ') # Remove newlines, strip whitespace 
                # Optional: Standardize common terms (e.g., 'Transaction Date' -> 'Date') 
                new_name = re.sub(r'transaction\s+date', 'Date', new_name, flags=re.IGNORECASE)
                new_name = re.sub(r'narration|particulars|details', 'Description', new_name, flags=re.IGNORECASE) # Added details
                new_name = re.sub(r'withdrawals?', 'Debit', new_name, flags=re.IGNORECASE)
                new_name = re.sub(r'deposits?', 'Credit', new_name, flags=re.IGNORECASE) 
                # Add more standardization rules as needed
            new_columns.append(new_name)

        clean_df.columns = new_columns

        # --- Stage 3: Remove completely empty columns --- 
        clean_df.dropna(axis=1, how='all', inplace=True)

        if self.debug:
            logger.debug(f"Cleaned columns: {clean_df.columns.tolist()}") 

        return clean_df

    def _identify_columns(self, df):
        """Identifies Date, Description, Amount (or Debit/Credit), Balance columns.""" 
        col_mapping = {}
        cols = list(df.columns)
        potential_matches = defaultdict(list)

        # --- Pass 1: Identify based on common column header keywords --- 
        keywords = { 
            'date': ['date', 'txn date'],
            'desc': ['description', 'details', 'narration', 'particulars', 'memo'],
            'amount': ['amount', 'value'],
            'debit': ['debit', 'withdrawal', 'payment', 'outgoing'],
            'credit': ['credit', 'deposit', 'payment in', 'incoming'],
            'balance': ['balance', 'running balance', 'acct balance'],
        } 

        assigned_cols = set()
        for standard_name, terms in keywords.items():
            for col in cols:
                 if col in assigned_cols: continue # Skip if already assigned 
                 col_lower = str(col).lower()
                 if any(term in col_lower for term in terms): 
                      potential_matches[standard_name].append(col)

        # Assign best match (often the first one found for each type) 
        for standard_name, matched_cols in potential_matches.items():
             if matched_cols:
                 best_col = matched_cols[0] # Simple: take the first match 
                 col_mapping[standard_name] = best_col
                 assigned_cols.add(best_col) 

        # --- Pass 2: Identify based on content if headers failed --- 
        # Check for Date if not found by header 
        if 'date' not in col_mapping:
            for col in cols: 
                 if col in assigned_cols: continue
                 sample = df[col].astype(str).dropna().head(5)
                 if sample.empty: continue
                 # Regex for common date formats 
                 date_pattern = r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}\s+[A-Za-z]{3})' 
                 if sum(sample.str.contains(date_pattern, regex=True)) >= 2:
                      col_mapping['date'] = col
                      assigned_cols.add(col)
                      break # Found date column

        # Check for Description if not found by header (look for longer text) 
        if 'desc' not in col_mapping:
             for col in cols:
                  if col in assigned_cols: continue
                  sample = df[col].astype(str).dropna().head(5)
                  if sample.empty: continue 
                  # Heuristic: contains letters and likely more than one word on average 
                  if sample.str.contains(r'[a-zA-Z]').all() and (sample.str.count(r'\s+').mean() >= 0.5):
                       col_mapping['desc'] = col
                       assigned_cols.add(col) 
                       break

        # Check for numeric columns (Amount/Debit/Credit/Balance) 
        numeric_cols = []
        for col in cols:
            if col in assigned_cols: continue
            sample = df[col].astype(str).dropna().head(5)
            if sample.empty: continue 
            # Regex for currency-like values (allows symbols, commas, decimals, negatives) 
            amount_pattern = r'^\s*\(?[$€£]?[\d,.-]+\)?\s*$'
            if sum(sample.str.match(amount_pattern)) >= 2: # At least 2 look like amounts 
                 numeric_cols.append(col)

        # Try to assign remaining numeric columns 
        remaining_numeric = [col for col in numeric_cols if col not in assigned_cols] 
        if remaining_numeric:
            # If separate debit/credit not found, maybe one is Amount? 
            if 'debit' not in col_mapping and 'credit' not in col_mapping and 'amount' not in col_mapping: 
                 # Heuristic: Does one have negative signs or parentheses? 
                 col_with_negatives = None 
                 for col in remaining_numeric:
                      sample = df[col].astype(str).dropna().head(10)
                      if sample.str.contains(r'[-()]').any():
                           col_with_negatives = col
                           break 
                 if col_with_negatives:
                      col_mapping['amount'] = col_with_negatives
                      assigned_cols.add(col_with_negatives)
                      remaining_numeric.remove(col_with_negatives) 
                 elif len(remaining_numeric) == 1: # Only one numeric column left? 
                      # Assume it's Amount 
                      col_mapping['amount'] = remaining_numeric[0]
                      assigned_cols.add(remaining_numeric[0])
                      remaining_numeric.pop(0)


            # If Amount OR Debit/Credit assigned, the last numeric is likely Balance 
            if ('amount' in col_mapping or ('debit' in col_mapping and 'credit' in col_mapping)) \
               and 'balance' not in col_mapping and len(remaining_numeric) >= 1: 
                 # Heuristic: Balance is often the rightmost numeric column 
                 col_mapping['balance'] = remaining_numeric[-1]
                 assigned_cols.add(remaining_numeric[-1]) 


        # --- Final check: Handle Debit/Credit if Amount wasn't identified --- 
        if 'amount' not in col_mapping and 'debit' in col_mapping and 'credit' in col_mapping:
             logger.info("Using separate Debit/Credit columns to derive Amount.") 
             # No actual modification needed here, the extraction loop will handle it 
             pass # The main extraction loop will use debit_col and credit_col 
        elif 'amount' not in col_mapping and ('debit' in col_mapping or 'credit' in col_mapping):
             # If only one of debit/credit found, treat it as 'amount' (might need sign adjustment) 
             single_flow_col = col_mapping.get('debit') or col_mapping.get('credit')
             if single_flow_col:
                 logger.warning(f"Only one flow column ('{single_flow_col}') found. Treating as 'amount'. Sign might be incorrect.")
                 col_mapping['amount'] = single_flow_col


        # Log warnings if essential columns are still missing 
        if 'date' not in col_mapping: logger.warning("Could not identify Date column.")
        if 'desc' not in col_mapping: logger.warning("Could not identify Description column.")
        if 'amount' not in col_mapping and ('debit' not in col_mapping or 'credit' not in col_mapping):
             logger.warning("Could not identify Amount (or Debit/Credit pair) column(s).") 

        return col_mapping

    def _parse_amount(self, amount_str):
        """Parse amount string (including currency, commas, negatives) to float.""" 
        if pd.isna(amount_str) or amount_str is None: return None
        amount_str = str(amount_str).strip()
        if not amount_str or amount_str.lower() in ['nan', 'none', '-']: return None

        # Remove currency symbols and thousand separators (commas) 
        # Keep decimal point and minus sign 
        cleaned = re.sub(r'[$\s€£,]', '', amount_str) if amount_str is not None else ''

        is_negative = False
        # Check for parentheses like (123.45) 
        if cleaned.startswith('(') and cleaned.endswith(')'):
            cleaned = cleaned[1:-1]
            is_negative = True
        # Check for trailing minus like 123.45- or trailing CR/DR 
        elif cleaned.endswith('-'):
            cleaned = cleaned[:-1]
            is_negative = True
        elif cleaned.endswith('CR'): # Credit Record often means positive balance/amount contextually 
             cleaned = cleaned[:-2]
             # is_negative = False # Explicitly positive or depends on context 
        elif cleaned.endswith('DR'): # Debit Record often means negative 
             cleaned = cleaned[:-2]
             is_negative = True

        # Check for leading minus (can happen even with parentheses sometimes) 
        if cleaned.startswith('-'):
             cleaned = cleaned[1:]
             is_negative = True 

        try:
            # Handle potential edge case where cleaned string is just "." or empty or not a string
            if not isinstance(cleaned, str) or cleaned is None or cleaned == '.' or cleaned.strip() == '':
                return None
            try:
                amount = float(cleaned)
            except Exception:
                return None
            return -amount if is_negative else amount
        except ValueError:
            if self.debug:
                logger.warning(f"Could not parse amount: '{amount_str}' -> '{cleaned}'") 
            return None # Return None if parsing fails 


    def _parse_date(self, date_str):
        """
        Parses a date string from the table, using statement period context. 
        Handles various formats like 'DD', 'DD Mon', 'Mon DD', 'MM/DD'. 
        Tracks the current month/year and handles year rollovers. 
        Returns date in 'YYYY-MM-DD' format or None if parsing fails. 
        """
        if not date_str or pd.isna(date_str): 
            return None

        date_str = str(date_str).strip()
        # Normalize common separators and remove extra spaces 
        date_str = re.sub(r'[/\-. ]+', ' ', date_str).strip()

        # Use current_parsing_year/month instance variables for context
        if not self.current_parsing_year or not self.current_parsing_month: 
             # If context is missing (shouldn't happen if init ran correctly)
             # Try to use start date as default context
             if self.start_year and self.start_month:
                 self.current_parsing_year = self.start_year
                 self.current_parsing_month = self.start_month
                 logger.warning("Parsing context missing, initialized from statement start date.")
             else:
                 logger.error("Cannot parse date accurately: Statement start date context is missing.") 
                 return None # Cannot proceed without context

        # --- Date parsing with statement context --- 
        parsed_dt = None
        day = None 
        month = None
        year = self.current_parsing_year # Assume current year initially

        # Attempt different parsing strategies 
        try:
            # Strategy 1: Format includes month (e.g., "15 Jan", "Jan 15", "01 15") 
            match_mon = re.match(r'(\d{1,2})\s+([A-Za-z]{3,})', date_str, re.IGNORECASE) # DD Mon
            if not match_mon: 
                 match_mon = re.match(r'([A-Za-z]{3,})\s+(\d{1,2})', date_str, re.IGNORECASE) # Mon DD

            if match_mon:
                 day_str, mon_str = match_mon.groups() if len(match_mon.groups()) == 2 else (match_mon.group(2), match_mon.group(1))
                 day = int(day_str) 
                 try: 
                      # Use datetime to parse month abbreviation robustly
                      month = datetime.strptime(mon_str, '%b').month if len(mon_str) == 3 else datetime.strptime(mon_str, '%B').month
                 except ValueError:
                      logger.warning(f"Could not parse month name: '{mon_str}' in '{date_str}'") 
                      return None

            # Strategy 2: Format is just day (e.g., "01", "5", "15") 
            elif re.fullmatch(r'\d{1,2}', date_str):
                 day = int(date_str)
                 month = self.current_parsing_month # Assume current month 

            # Strategy 3: Format like MM DD or MM/DD (less common in statements typically) 
            elif re.fullmatch(r'(\d{1,2})\s+(\d{1,2})', date_str):
                mon_str, day_str = date_str.split()
                day = int(day_str)
                month = int(mon_str) 

            # Strategy 4: Full date format (e.g., YYYY-MM-DD, DD/MM/YYYY) - try pandas first
            else:
                try:
                    # Let pandas try its best first
                    parsed_dt = pd.to_datetime(date_str)
                    day = parsed_dt.day
                    month = parsed_dt.month
                    year = parsed_dt.year # Get year directly if available
                except Exception:
                    logger.warning(f"Could not parse date string '{date_str}' with pandas or specific patterns.")
                    return None


            # If parsing strategy yielded day and month (or full dt from pandas): 
            if day is not None and month is not None:
                # --- Contextual Adjustments ---
                # 1. Handle month change: If the new month is numerically less than the 
                #    current tracked month, it likely means the year has rolled over. 
                #    (e.g., current is Dec (12), new is Jan (1) -> increment year) 
                # Only adjust year if year wasn't explicit in parsed date
                if parsed_dt is None or parsed_dt.year == 1900: # Pandas default year can be 1900
                    if month < self.current_parsing_month and self.current_parsing_month == 12 and month == 1:
                        year += 1
                        logger.debug(f"Year rollover detected: Month changed from {self.current_parsing_month} to {month}. New year: {year}") 
                    elif month > self.current_parsing_month: 
                        # Month simply advanced within the same year (or started)
                        pass # Year stays the same
                    # Else (month == current_parsing_month): Year stays the same

                # 2. Validate day/month/year combination 
                if not (1 <= month <= 12):
                     logger.warning(f"Invalid month ({month}) parsed from '{date_str}'")
                     return None
                max_days = calendar.monthrange(year, month)[1] 
                if not (1 <= day <= max_days):
                     logger.warning(f"Invalid day ({day}) for month {month}/{year} parsed from '{date_str}'")
                     return None


                # 3. Construct datetime object if not already parsed by pandas 
                if parsed_dt is None:
                    parsed_dt = datetime(year, month, day) 
                else: # Ensure year is correct even if pandas parsed it
                    parsed_dt = parsed_dt.replace(year=year)

                # 4. Update current parsing context 
                self.current_parsing_month = month
                self.current_parsing_year = year

                return parsed_dt.strftime('%Y-%m-%d')

            else: # This case should be less likely now with pandas fallback
                 logger.warning(f"Could not parse date string: '{date_str}' using known patterns.") 
                 return None

        except ValueError as e:
            logger.warning(f"Error parsing date string '{date_str}': {e}") 
            return None
        except Exception as e_gen: # Catch any other unexpected errors 
            logger.error(f"Unexpected error parsing date '{date_str}': {e_gen}", exc_info=self.debug)
            return None

    def _format_date_for_excel(self, date_input):
        """Format date object or string into a more detailed format (DD-Month-YYYY).""" 
        if pd.isna(date_input) or date_input is None: return None

        date_obj = None
        if isinstance(date_input, datetime):
            date_obj = date_input
        elif isinstance(date_input, str):
            date_str = date_input.strip()
            if not date_str or date_str.lower() in ['nan', 'none', 'date']: return None 
            try:
                # Try standard format first 
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                # Try other common formats 
                for fmt in [
                    '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d',
                    '%d/%m/%y', '%d-%m-%y', '%Y-%m-%d',
                    '%m/%d/%Y', '%m-%d-%Y',
                    '%d-%b-%Y', '%d %b %Y',  # For dates like 01-Jan-2024 or 01 Jan 2024 
                    '%b %d %Y',              # For dates like Jan 01 2024
                    '%d-%B-%Y', '%d %B %Y',  # For dates like 01-January-2024 or 01 January 2024 
                    '%B %d %Y',               # For dates like January 01 2024 
                    '%d %b, %Y',              # For dates like 01 Feb, 2024 
                    '%d %B, %Y',              # For dates like 01 February, 2024 
                    '%b %d, %Y',              # For dates like Feb 01, 2024
                    '%B %d, %Y',              # For dates like February 01, 2024 
                ]:
                    try:
                        parsed = datetime.strptime(date_str, fmt)
                        # Handle 2-digit years 
                        if parsed.year < 100:
                            if parsed.year < 50:
                                parsed = parsed.replace(year=parsed.year + 2000) 
                            else:
                                parsed = parsed.replace(year=parsed.year + 1900) 
                        date_obj = parsed
                        break # Exit loop on successful parse 
                    except ValueError:
                        continue # Try next format 

                # Regex fallback if formats fail 
                if date_obj is None:
                    try:
                        # Match patterns like DD-MM-YYYY, DD/MM/YYYY, etc. 
                        match = re.match(r'(\d{1,2})[-/.\s](\d{1,2})[-/.\s](\d{2,4})', date_str)
                        if match:
                            day, month, year = match.groups() 
                            day = int(day); month = int(month); year = int(year) 
                            if year < 100: # Handle 2-digit year 
                                year = year + 2000 if year < 50 else year + 1900
                            try: 
                                date_obj = datetime(year, month, day)
                            except ValueError: 
                                pass 
                    except Exception: 
                        pass
        else: # Input is not datetime or string
            logger.warning(f"Invalid input type for date formatting: {type(date_input)}")
            return None

        # If we have a valid date object, format it
        if date_obj:
             try:
                 return date_obj.strftime('%d-%B-%Y') # e.g., "01-January-2024" 
             except ValueError: # Handle potential issues with very old dates if necessary
                 logger.warning(f"Could not format date object: {date_obj}")
                 return str(date_input) # Return original input as string
        else:
            logger.warning(f"Failed to parse or format date: {date_input}")
            return str(date_input) # Return original input as string if all fails

    def generate_account_summary(self):
        """Generate a dictionary summarizing account activity.""" 
        summary = {} 

        # --- Summary Info Section --- 
        summary['Summary Info'] = {
            'Name of the Account Holder': self.account_info.get('owner_name', 'N/A'),
            'Address': self.account_info.get('address', 'N/A'), # Assuming you might add address extraction
            'Email': self.account_info.get('email', 'N/A'), # Assuming you might add email extraction
            'PAN': self.account_info.get('pan', 'N/A'), # Assuming you might add PAN extraction
            'Mobile Number': self.account_info.get('mobile', 'N/A'), 
            'Name of the Bank': self.account_info.get('bank_name', 'N/A'),
            'Account Number': self.account_info.get('account_number', 'N/A'),
            'Account Type': self.account_info.get('account_type', 'N/A'),
            'EMI Amount': 'N/A', # Placeholder, calculate from analysis if needed
            'Transaction ID': self.unique_id 
        }

        # --- Monthwise Details Section --- 
        # Use analysis_results['monthly_summary'] which has opening balance
        monthly_summary_data = self.analysis_results.get('monthly_summary', {})
        if monthly_summary_data: 
            monthly_details = {}
            all_balances = [] # For average calculation
            for month_key in sorted(monthly_summary_data.keys()): 
                try:
                    data = monthly_summary_data[month_key]
                    year, month = map(int, month_key.split('-')) 
                    month_name = calendar.month_name[month]
                    short_month = f"{month_name[:3]}-{str(year)[2:]}"  # e.g., "Jun-24" 

                    # Get opening and closing balance for the month 
                    # Use the balances calculated in _calculate_monthly_balances
                    opening_balance = data.get('opening_balance') # Balance before first txn of month
                    balance_on_5th = data.get('balance_after_5th') # Balance on/after 5th

                    # Find closing balance for the month (last transaction's balance)
                    closing_balance = None
                    transactions_this_month = [
                        t['balance'] for t in self.transactions
                        if t.get('date') and t['date'].startswith(month_key) and t.get('balance') is not None
                    ]
                    if transactions_this_month:
                        closing_balance = transactions_this_month[-1]
                        all_balances.append(closing_balance) # Add month-end balance for averaging

                    monthly_details[short_month] = { 
                        'Opening Balance': f"{opening_balance:.2f}" if opening_balance is not None else 'N/A',
                        'Closing Balance': f"{closing_balance:.2f}" if closing_balance is not None else 'N/A',
                        'Balance as on 5th': f"{balance_on_5th:.2f}" if balance_on_5th is not None else 'N/A' 
                    }
                except Exception as e:
                    logger.error(f"Error processing month {month_key} for summary: {e}") 
                    continue 

            summary['Monthwise Details'] = monthly_details 

            # --- Calculate Averages (Simplified Example) --- 
            # Requires more transactions over longer periods for meaningful averages
            num_months = len(all_balances)
            avg_3m = avg_6m = avg_12m = 'NA'
            if num_months >= 3: avg_3m = f"{(sum(all_balances[-3:]) / 3):.2f}"
            if num_months >= 6: avg_6m = f"{(sum(all_balances[-6:]) / 6):.2f}"
            if num_months >= 12: avg_12m = f"{(sum(all_balances[-12:]) / 12):.2f}"

            # Calculate transaction averages (requires tracking monthly counts/amounts)
            # This part needs more detailed implementation within transaction processing or analysis
            avg_credit_count_6m = 'NA'
            avg_credit_amount_6m = 'NA'
            avg_debit_count_6m = 'NA'

        else:
            summary['Monthwise Details'] = {} 
            avg_3m = avg_6m = avg_12m = 'NA'
            avg_credit_count_6m = 'NA'
            avg_credit_amount_6m = 'NA'
            avg_debit_count_6m = 'NA'

        # --- Overall Summary Section --- 
        summary['Overall Summary'] = {
            'Average Balance of 3 Months': avg_3m,
            'Average Balance of 6 Months': avg_6m,
            'Average Balance of 12 Months': avg_12m,
            'Average of No.of Credit Transaction of 6 months': avg_credit_count_6m, 
            'Average of Amount of Credit Transaction of 6 Months': avg_credit_amount_6m,
            'Average No.of Debit Transaction of 6 months': avg_debit_count_6m 
        }

        # Add identified EMIs/Salary to summary if desired
        summary['Analysis'] = {
            'Identified EMIs': self.analysis_results.get('identified_emis', []),
            'Identified Salary': self.analysis_results.get('identified_salary', [])
        }

        return summary


    def export_to_excel(self, output_path): 
        """
        Export extracted transactions, account info, and analysis results to an Excel file. 
        """ 
        try:
            wb = openpyxl.Workbook() 
            # Remove default sheet if needed, or rename it 
            if "Sheet" in wb.sheetnames:
                default_sheet = wb["Sheet"]
                wb.remove(default_sheet)

            header_font = Font(bold=True) 

            # --- 1. All Transactions Sheet --- 
            all_trans_ws = wb.create_sheet("All Transactions")
            all_trans_ws.column_dimensions['A'].width = 18 # Date (DD-Month-YYYY) 
            all_trans_ws.column_dimensions['B'].width = 50 # Description
            all_trans_ws.column_dimensions['C'].width = 15 # Debit
            all_trans_ws.column_dimensions['D'].width = 15 # Credit 
            all_trans_ws.column_dimensions['E'].width = 15 # Balance 

            # Add bank name, account number, and transaction ID at the top 
            current_header_row = 1
            if self.account_info.get('bank_name'):
                all_trans_ws.cell(row=current_header_row, column=1, value="Bank Name:").font = header_font
                all_trans_ws.cell(row=current_header_row, column=2, value=self.account_info['bank_name'])
                current_header_row += 1 

            # Add account number prominently 
            if self.account_info.get('account_number'):
                all_trans_ws.cell(row=current_header_row, column=1, value="Account Number:").font = header_font
                cell_acc_num = all_trans_ws.cell(row=current_header_row, column=2, value=self.account_info['account_number'])
                cell_acc_num.font = Font(bold=True, size=12) # Make account number bold and slightly larger 
                current_header_row += 1

            all_trans_ws.cell(row=current_header_row, column=1, value="Transaction ID:").font = header_font 
            all_trans_ws.cell(row=current_header_row, column=2, value=self.unique_id)
            current_header_row += 2 # Add a blank row before headers

            headers_all = ['Date', 'Description', 'Category', 'Debit', 'Credit', 'Balance'] 
            for col, header in enumerate(headers_all, 1):
                cell = all_trans_ws.cell(row=current_header_row, column=col)
                cell.value = header
                cell.font = header_font 

            # Start data rows after the headers
            data_start_row = current_header_row + 1
            for row_idx, trans in enumerate(self.transactions, data_start_row):
                formatted_date = self._format_date_for_excel(trans.get('date')) 
                amount = trans.get('amount', 0) 
                debit_val = abs(amount) if amount is not None and amount < 0 else None
                credit_val = amount if amount is not None and amount > 0 else None
                balance_val = trans.get('balance')
                category_val = trans.get('category', 'Other')
                all_trans_ws.cell(row=row_idx, column=1, value=formatted_date) 
                all_trans_ws.cell(row=row_idx, column=2, value=trans.get('description'))
                all_trans_ws.cell(row=row_idx, column=3, value=category_val)
                cell_debit = all_trans_ws.cell(row=row_idx, column=4, value=debit_val)
                cell_credit = all_trans_ws.cell(row=row_idx, column=5, value=credit_val)
                cell_balance = all_trans_ws.cell(row=row_idx, column=6, value=balance_val)
                if debit_val is not None: cell_debit.number_format = '#,##0.00'
                if credit_val is not None: cell_credit.number_format = '#,##0.00'
                if balance_val is not None: cell_balance.number_format = '#,##0.00'


            # --- 2. High Value Transactions Sheet --- 
            high_value_ws = wb.create_sheet("High Value Transactions") 
            high_value_ws.column_dimensions['A'].width = 18 # Date 
            high_value_ws.column_dimensions['B'].width = 50 # Description
            high_value_ws.column_dimensions['C'].width = 15 # Debit
            high_value_ws.column_dimensions['D'].width = 15 # Credit
            high_value_ws.column_dimensions['E'].width = 15 # Balance

            headers_high = ['Date', 'Description', 'Category', 'Debit', 'Credit', 'Balance'] 
            for col, header in enumerate(headers_high, 1): 
                cell = high_value_ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font

            for row_idx, trans in enumerate(self.high_value_transactions, 2): # Use self.high_value_transactions 
                 formatted_date = self._format_date_for_excel(trans.get('date')) 
                 amount = trans.get('amount', 0)
                 debit_val = abs(amount) if amount is not None and amount < 0 else None
                 credit_val = amount if amount is not None and amount > 0 else None
                 balance_val = trans.get('balance')
                 category_val = trans.get('category', 'Other')
                 high_value_ws.cell(row=row_idx, column=1, value=formatted_date) 
                 high_value_ws.cell(row=row_idx, column=2, value=trans.get('description'))
                 high_value_ws.cell(row=row_idx, column=3, value=category_val)
                 cell_debit_hv = high_value_ws.cell(row=row_idx, column=4, value=debit_val)
                 cell_credit_hv = high_value_ws.cell(row=row_idx, column=5, value=credit_val)
                 cell_balance_hv = high_value_ws.cell(row=row_idx, column=6, value=balance_val)
                 if debit_val is not None: cell_debit_hv.number_format = '#,##0.00'
                 if credit_val is not None: cell_credit_hv.number_format = '#,##0.00'
                 if balance_val is not None: cell_balance_hv.number_format = '#,##0.00'


            # --- 3. Analysis Sheet --- 
            analysis_ws = wb.create_sheet("Analysis") 
            analysis_ws.column_dimensions['A'].width = 35
            analysis_ws.column_dimensions['B'].width = 25
            analysis_ws.column_dimensions['C'].width = 18
            analysis_ws.column_dimensions['D'].width = 18
            analysis_ws.column_dimensions['E'].width = 30 # Wider for date ranges

            current_row = 1 

            # --- Header Section ---
            analysis_ws.cell(row=current_row, column=1, value="Extraction ID:").font = header_font
            analysis_ws.cell(row=current_row, column=2, value=self.unique_id)
            current_row += 1
            analysis_ws.cell(row=current_row, column=1, value="Start Date:").font = header_font
            analysis_ws.cell(row=current_row, column=2, value=self.start_date.strftime('%d-%B-%Y'))
            current_row += 1
            # Always use the bank name as selected by the user (self.account_info['bank_name'])
            bank_name_val = self.account_info.get('bank_name', 'N/A')
            analysis_ws.cell(row=current_row, column=1, value="Bank Name:").font = header_font
            analysis_ws.cell(row=current_row, column=2, value=bank_name_val)
            current_row += 2
            # --- Account Information Section ---
            analysis_ws.cell(row=current_row, column=1, value="Account Information").font = header_font
            current_row += 1
            analysis_ws.cell(row=current_row, column=1, value="Account Number").font = header_font
            analysis_ws.cell(row=current_row, column=2, value=self.account_info.get('account_number', 'N/A'))
            current_row += 1
            if self.account_info.get('tax_id'):
                analysis_ws.cell(row=current_row, column=1, value="Tax ID").font = header_font
                analysis_ws.cell(row=current_row, column=2, value=self.account_info.get('tax_id'))
                current_row += 1
            # Only show one Opening Balance, and match it to the first month in Monthly Summary if available
            monthly_summary_data = self.analysis_results.get('monthly_summary', {})
            first_month = None
            first_month_ob = None
            if monthly_summary_data:
                first_month = sorted(monthly_summary_data.keys())[0]
                first_month_ob = monthly_summary_data[first_month].get('opening_balance')
            if first_month_ob is not None:
                try:
                    ob_val = float(first_month_ob)
                    analysis_ws.cell(row=current_row, column=1, value="Opening Balance").font = header_font
                    analysis_ws.cell(row=current_row, column=2, value=ob_val)
                    current_row += 1
                except Exception:
                    pass
            if self.account_info.get('closing_balance') is not None:
                try:
                    cb_val = float(self.account_info.get('closing_balance'))
                    analysis_ws.cell(row=current_row, column=1, value="Closing Balance").font = header_font
                    analysis_ws.cell(row=current_row, column=2, value=cb_val)
                    current_row += 1
                except Exception:
                    pass
            current_row += 1
            # --- Monthly Summary Section ---
            analysis_ws.cell(row=current_row, column=1, value="Monthly Summary").font = header_font
            current_row += 1
            headers_monthly = [
                "Month",
                "Opening Balance",
                "Balance After 5th",
                "Balance After 15th",
                "Balance After 25th"
            ]
            for col, header in enumerate(headers_monthly, 1):
                cell = analysis_ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
            current_row += 1
            monthly_summary_data = self.analysis_results.get('monthly_summary', {})
            if monthly_summary_data:
                for month, data in sorted(monthly_summary_data.items()):
                    opening_bal = data.get('opening_balance')
                    bal_after_5 = data.get('balance_after_5th')
                    bal_after_15 = data.get('balance_after_15th')
                    bal_after_25 = data.get('balance_after_25th')
                    analysis_ws.cell(row=current_row, column=1, value=month)
                    cell_ob = analysis_ws.cell(row=current_row, column=2, value=opening_bal)
                    cell_b5 = analysis_ws.cell(row=current_row, column=3, value=bal_after_5)
                    cell_b15 = analysis_ws.cell(row=current_row, column=4, value=bal_after_15)
                    cell_b25 = analysis_ws.cell(row=current_row, column=5, value=bal_after_25)
                    if opening_bal is not None: cell_ob.number_format = '#,##0.00'
                    if bal_after_5 is not None: cell_b5.number_format = '#,##0.00'
                    if bal_after_15 is not None: cell_b15.number_format = '#,##0.00'
                    if bal_after_25 is not None: cell_b25.number_format = '#,##0.00'
                    current_row += 1
            else:
                analysis_ws.cell(row=current_row, column=1, value="No monthly summary data available.")
                current_row += 1
            current_row += 1
            # --- Identified EMIs/Loans Section --- 
            analysis_ws.cell(row=current_row, column=1, value="Identified EMIs/Recurring Debits").font = Font(bold=True)
            current_row += 1
            emis_data = self.analysis_results.get('identified_emis', []) 
            if emis_data: 
                analysis_ws.cell(row=current_row, column=1, value="Description Pattern").font = Font(bold=True)
                analysis_ws.cell(row=current_row, column=2, value="Estimated Amount").font = Font(bold=True)
                analysis_ws.cell(row=current_row, column=3, value="Occurrences").font = Font(bold=True)
                analysis_ws.cell(row=current_row, column=4, value="First Date").font = Font(bold=True)
                analysis_ws.cell(row=current_row, column=5, value="Last Date").font = Font(bold=True) 
                current_row += 1
                for emi in emis_data:
                    est_amount = emi.get('estimated_amount')
                    first_date = self._format_date_for_excel(emi.get('first_occurrence_date'))
                    last_date = self._format_date_for_excel(emi.get('last_occurrence_date')) 

                    analysis_ws.cell(row=current_row, column=1, value=emi.get('description_pattern'))
                    cell_emi_amt = analysis_ws.cell(row=current_row, column=2, value=est_amount)
                    analysis_ws.cell(row=current_row, column=3, value=emi.get('occurrences'))
                    analysis_ws.cell(row=current_row, column=4, value=first_date) 
                    analysis_ws.cell(row=current_row, column=5, value=last_date)

                    if est_amount is not None: cell_emi_amt.number_format = '#,##0.00'
                    current_row += 1
            else:
                 analysis_ws.cell(row=current_row, column=1, value="None Identified") 
                 current_row += 1
            current_row += 1 # Add gap


            # --- Identified Salary Section --- 
            analysis_ws.cell(row=current_row, column=1, value="Identified Salary/Regular Credits").font = Font(bold=True)
            current_row += 1
            salary_data = self.analysis_results.get('identified_salary', []) 
            if salary_data: 
                analysis_ws.cell(row=current_row, column=1, value="Identification Type").font = Font(bold=True)
                analysis_ws.cell(row=current_row, column=2, value="Description Sample").font = Font(bold=True)
                analysis_ws.cell(row=current_row, column=3, value="Estimated Amount").font = Font(bold=True)
                analysis_ws.cell(row=current_row, column=4, value="Occurrences").font = Font(bold=True) 
                analysis_ws.cell(row=current_row, column=5, value="Date Range").font = Font(bold=True)
                current_row += 1
                for sal in salary_data:
                    est_amount = sal.get('estimated_amount') 
                    first_date = self._format_date_for_excel(sal.get('first_occurrence_date')) 
                    last_date = self._format_date_for_excel(sal.get('last_occurrence_date'))

                    analysis_ws.cell(row=current_row, column=1, value=sal.get('type'))
                    analysis_ws.cell(row=current_row, column=2, value=sal.get('description_sample'))
                    cell_sal_amt = analysis_ws.cell(row=current_row, column=3, value=est_amount)
                    analysis_ws.cell(row=current_row, column=4, value=sal.get('occurrences')) 
                    analysis_ws.cell(row=current_row, column=5, value=f"{first_date} to {last_date}") # Combined date range

                    if est_amount is not None: cell_sal_amt.number_format = '#,##0.00'
                    current_row += 1
            else: 
                analysis_ws.cell(row=current_row, column=1, value="None Identified")
                current_row += 1
            # End of Analysis Sheet content


            # --- Save Workbook --- 
            wb.save(output_path)
            logger.info(f"Excel file with analysis saved successfully at: {output_path}") 

        # Use ImportError specific to openpyxl if possible, or broader ImportError 
        except ImportError:
            logger.error("The 'openpyxl' library is required for Excel export. Please install it (pip install openpyxl).") 
            raise # Re-raise the error so the calling code knows export failed 
        except Exception as e:
            logger.error(f"Error exporting data to Excel: {e}", exc_info=True) 
            # Optionally re-raise the exception if the calling code needs to handle it
            # raise e


    def to_xml(self, output_path=None): 
        """Export extracted data to a structured XML file.""" 
        if not output_path:
            base_name = os.path.splitext(os.path.basename(self.pdf_path))[0]
            output_path = f"{base_name}_{self.unique_id}_summary.xml"

        logger.info(f"Exporting data to XML: {output_path}")

        try:
            # --- Create Root Element --- 
            root = ET.Element('BankStatement') 
            root.set('transactionId', self.unique_id)
            root.set('sourcePdf', os.path.basename(self.pdf_path))
            root.set('generatedTimestamp', datetime.now().isoformat())
            if self.account_info.get('bank_name'):
                root.set('bankName', self.account_info['bank_name'])

            # --- Account Info Element --- 
            account_info_elem = ET.SubElement(root, 'AccountInformation') 
            for key, value in self.account_info.items():
                # Clean key for XML tag name (basic) 
                tag_name = re.sub(r'\s+', '_', key).strip()
                tag_name = re.sub(r'[^a-zA-Z0-9_.-]', '', tag_name) # Allow dot and hyphen
                # Ensure starts with letter or underscore
                if not tag_name or not (tag_name[0].isalpha() or tag_name[0] == '_'):
                    tag_name = '_' + tag_name # Prepend underscore if invalid start

                if tag_name: # Ensure tag name is not empty after cleaning 
                    ET.SubElement(account_info_elem, tag_name).text = str(value) if value is not None else ""

            # --- Account Summary Element --- 
            summary = self.generate_account_summary()
            summary_elem = ET.SubElement(root, 'AccountSummary')
            for section, data in summary.items(): 
                section_tag = re.sub(r'\s+', '', section) # e.g., SummaryInfo, MonthwiseDetails
                section_elem = ET.SubElement(summary_elem, section_tag)

                if isinstance(data, dict):
                    for key, value in data.items():
                        key_tag = re.sub(r'\s+', '', key) # e.g., NameOfTheAccountHolder
                        key_tag = re.sub(r'[^a-zA-Z0-9_.-]', '', key_tag)
                        if not key_tag or not (key_tag[0].isalpha() or key_tag[0] == '_'):
                           key_tag = '_' + key_tag

                        if isinstance(value, dict): # Handle nested dict like Monthwise Details
                            sub_section_elem = ET.SubElement(section_elem, key_tag)
                            for sub_key, sub_value in value.items():
                                sub_key_tag = re.sub(r'\s+', '', sub_key)
                                sub_key_tag = re.sub(r'[^a-zA-Z0-9_.-]', '', sub_key_tag)
                                if not sub_key_tag or not (sub_key_tag[0].isalpha() or sub_key_tag[0] == '_'):
                                   sub_key_tag = '_' + sub_key_tag
                                ET.SubElement(sub_section_elem, sub_key_tag).text = str(sub_value) if sub_value is not None else ""
                        elif isinstance(value, list): # Handle lists like EMIs/Salary
                            list_elem = ET.SubElement(section_elem, key_tag)
                            for item in value:
                                if isinstance(item, dict):
                                    item_tag = "Item" # Generic tag for list items
                                    item_elem = ET.SubElement(list_elem, item_tag)
                                    for item_key, item_val in item.items():
                                        item_key_tag = re.sub(r'\s+', '', item_key)
                                        item_key_tag = re.sub(r'[^a-zA-Z0-9_.-]', '', item_key_tag)
                                        if not item_key_tag or not (item_key_tag[0].isalpha() or item_key_tag[0] == '_'):
                                           item_key_tag = '_' + item_key_tag
                                        ET.SubElement(item_elem, item_key_tag).text = str(item_val) if item_val is not None else ""
                        else: # Simple key-value pair
                             ET.SubElement(section_elem, key_tag).text = str(value) if value is not None else ""


            # --- Monthly Balances Element (Now included in Summary, keep if separate needed) --- 
            # monthly_balances_elem = ET.SubElement(root, 'MonthlyBalancesFifth') 
            # monthly_summary_data = self.analysis_results.get('monthly_summary', {})
            # if monthly_summary_data:
            #      for month_key in sorted(monthly_summary_data.keys()):
            #           data = monthly_summary_data[month_key]
            #           month_elem = ET.SubElement(monthly_balances_elem, 'MonthBalance') 
            #           month_elem.set('monthKey', month_key) 
            #           ET.SubElement(month_elem, 'opening_balance_derived_from_date').text = self._format_date_for_excel(data.get('opening_balance_derived_from_date', 'N/A'))
            #           ET.SubElement(month_elem, 'balance_after_5th_date').text = self._format_date_for_excel(data.get('balance_after_5th_date', 'N/A'))
            #           opening_bal_val = data.get('opening_balance')
            #           bal_after_5_val = data.get('balance_after_5th')
            #           ET.SubElement(month_elem, 'opening_balance').text = f"{opening_bal_val:.2f}" if opening_bal_val is not None else 'N/A'
            #           ET.SubElement(month_elem, 'balance_after_5th').text = f"{bal_after_5_val:.2f}" if bal_after_5_val is not None else 'N/A' 


            # --- High Value Transactions Element --- 
            high_value_elem = ET.SubElement(root, 'HighValueTransactions')
            high_value_elem.set('threshold', str(self.high_value_threshold)) 
            if self.high_value_transactions:
                 for tx in self.high_value_transactions:
                      tx_elem = ET.SubElement(high_value_elem, 'Transaction') 
                      tx_elem.set('id', tx.get('transaction_id', ''))
                      for key, value in tx.items():
                           if key == 'is_high_value': continue # Skip flag in XML details 
                           # Clean key for XML tag name (reuse logic) 
                           tag_name = re.sub(r'\s+', '_', key).strip()
                           tag_name = re.sub(r'[^a-zA-Z0-9_.-]', '', tag_name)
                           if not tag_name or not (tag_name[0].isalpha() or tag_name[0] == '_'):
                               tag_name = '_' + tag_name

                           if tag_name: 
                                # Format numbers 
                                if isinstance(value, (int, float)):
                                     text_value = f"{value:.2f}" 
                                elif isinstance(value, datetime):
                                     text_value = value.strftime('%Y-%m-%d') # Format date consistently
                                else: 
                                     text_value = str(value) if value is not None else ""
                                ET.SubElement(tx_elem, tag_name).text = text_value


            # --- All Transactions Element --- 
            transactions_elem = ET.SubElement(root, 'AllTransactions')
            if self.transactions:
                 for tx in self.transactions:
                      tx_elem = ET.SubElement(transactions_elem, 'Transaction')
                      tx_elem.set('id', tx.get('transaction_id', '')) 
                      for key, value in tx.items():
                           if key == 'is_high_value': continue # Skip flag 
                           # Clean key for XML tag name (reuse logic) 
                           tag_name = re.sub(r'\s+', '_', key).strip()
                           tag_name = re.sub(r'[^a-zA-Z0-9_.-]', '', tag_name)
                           if not tag_name or not (tag_name[0].isalpha() or tag_name[0] == '_'):
                               tag_name = '_' + tag_name

                           if tag_name: 
                                # Format numbers and dates 
                                if isinstance(value, (int, float)): 
                                     text_value = f"{value:.2f}"
                                elif isinstance(value, datetime): 
                                     text_value = value.strftime('%Y-%m-%d')
                                else:
                                     text_value = str(value) if value is not None else "" 
                                ET.SubElement(tx_elem, tag_name).text = text_value

            # --- Convert ET to string with pretty printing --- 
            rough_string = ET.tostring(root, 'utf-8')
            reparsed = xml.dom.minidom.parseString(rough_string)
            pretty_xml_as_string = reparsed.toprettyxml(indent="  ", encoding='utf-8') # Ensure utf-8 output

            # --- Write to file --- 
            with open(output_path, "wb") as f: # Use 'wb' for bytes output from toprettyxml
                 f.write(pretty_xml_as_string) 

            logger.info(f"Successfully exported data to {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Failed to export data to XML: {e}", exc_info=self.debug) 
            raise # Re-raise the exception

    def _categorize_transaction(self, desc, amount):
        desc_lower = desc.lower()
        # Special handling for NACH
        if 'nach' in desc_lower:
            if any(k in desc_lower for k in ['loan', 'emi', 'instalment', 'installment']):
                return 'EMI'
            if any(k in desc_lower for k in ['sip', 'mutual fund', 'investment', 'pms', 'insurance']):
                return 'Investment'
            return 'NACH'
        if any(k in desc_lower for k in ['upi', 'gpay', 'google pay', 'phonepe', 'paytm']):
            return 'UPI'
        if any(k in desc_lower for k in ['neft']):
            return 'NEFT'
        if any(k in desc_lower for k in ['imps']):
            return 'IMPS'
        if any(k in desc_lower for k in ['atm', 'cash withdrawal']):
            return 'ATM'
        if any(k in desc_lower for k in ['cheque', 'chq', 'cq no', 'chq no']):
            return 'Cheque'
        if any(k in desc_lower for k in ['debit card', 'credit card', 'pos', 'card']):
            return 'Card'
        if any(k in desc_lower for k in ['salary', 'payroll', 'wages', 'remuneration', 'stipend']):
            return 'Salary'
        if any(k in desc_lower for k in ['emi', 'loan', 'instalment', 'installment']):
            return 'EMI'
        if any(k in desc_lower for k in ['interest', 'int.']):
            return 'Interest'
        if any(k in desc_lower for k in ['fee', 'charge', 'penalty', 'fine']):
            return 'Fee'
        if any(k in desc_lower for k in ['refund', 'reversal', 'chargeback']):
            return 'Refund'
        if any(k in desc_lower for k in ['transfer', 'to ', 'from ']):
            return 'Transfer'
        if any(k in desc_lower for k in ['cash deposit', 'cash']):
            return 'Cash'
        if amount > 0 and any(k in desc_lower for k in ['reversal', 'refund']):
            return 'Refund'
        return 'Other'


# --- Command Line Interface (Example Usage) ---
if __name__ == '__main__': 
    parser = argparse.ArgumentParser(description='Extract bank statement data from PDF to Excel/XML.') 
    parser.add_argument('pdf_path', help='Path to the bank statement PDF file.')
    # Add required date arguments for CLI usage
    parser.add_argument('start_date', help='Start month/year (MM/YYYY).')
    parser.add_argument('end_date', help='End month/year (MM/YYYY).')
    parser.add_argument('-p', '--password', help='Password for encrypted PDF.', default=None)
    parser.add_argument('-o', '--output-dir', help='Directory to save output files.', default='outputs') # Changed default
    parser.add_argument('-f', '--output-format', help='Output format: excel, xml, or both.', default='both', choices=['excel', 'xml', 'both'])
    parser.add_argument('--debug', help='Enable debug logging and save intermediate files.', action='store_true')
    parser.add_argument('--threshold', help='Threshold for high-value transactions.', type=float, default=1000)
    parser.add_argument('--bank', help='Name of the bank (optional).', default=None) # Add bank name arg

    args = parser.parse_args()

    # Ensure output directory exists 
    os.makedirs(args.output_dir, exist_ok=True)

    try: 
        # Initialize extractor
        extractor = BankStatementExtractor(
            pdf_path=args.pdf_path,
            start_date_str=args.start_date, # Pass dates
            end_date_str=args.end_date,
            password=args.password,
            debug=args.debug,
            high_value_threshold=args.threshold,
            bank_name=args.bank # Pass bank name
        )

        # Run extraction 
        print("Starting extraction process...")
        extractor.extract_all() # This runs text, account info, and transaction extraction 

        print(f"Extraction complete for ID: {extractor.unique_id}") 
        print(f"Found {len(extractor.transactions)} transactions.")
        print(f"Account info keys extracted: {', '.join(extractor.account_info.keys())}")

        # Generate and print summary to console 
        summary = extractor.generate_account_summary()
        print("\n--- Account Summary ---")
        # Print summary sections more nicely
        for section, data in summary.items():
            print(f"\n** {section.replace('_', ' ').title()} **")
            if isinstance(data, dict):
                for key, value in data.items():
                    if isinstance(value, dict): # Handle Monthwise Details
                        print(f"  {key}:")
                        for sub_key, sub_value in value.items():
                            print(f"    {sub_key.replace('_', ' ').title()}: {sub_value}")
                    elif isinstance(value, list): # Handle Analysis lists
                         print(f"  {key.replace('_', ' ').title()}:")
                         if value:
                            for item in value: print(f"    - {item}")
                         else: print("    None")
                    else: # Simple key-value
                        print(f"  {key.replace('_', ' ').title()}: {value}") 
            else:
                print(f"  {data}") # Should not happen with current structure
        print("-----------------------\n")


        # Export to selected formats 
        if args.output_format in ['excel', 'both']:
            base_name = os.path.splitext(os.path.basename(args.pdf_path))[0]
            excel_path = os.path.join(args.output_dir, f"{base_name}_{extractor.unique_id}_summary.xlsx") 
            extractor.export_to_excel(excel_path)
            print(f"Exported to Excel: {excel_path}")

        if args.output_format in ['xml', 'both']:
            base_name = os.path.splitext(os.path.basename(args.pdf_path))[0]
            xml_path = os.path.join(args.output_dir, f"{base_name}_{extractor.unique_id}_summary.xml")
            xml_file = extractor.to_xml(xml_path)
            print(f"Exported to XML: {xml_file}") 

    except FileNotFoundError as fnf:
         print(f"Error: Input PDF file not found at '{args.pdf_path}'") 
    except ValueError as ve: # Password errors etc. 
         print(f"Error: {ve}")
         if "password" in str(ve).lower():
             print("Hint: If the PDF is encrypted, use the --password option.")
    except ImportError:
         print("Error: Missing required library. Ensure 'pandas', 'openpyxl', 'PyPDF2', 'tabula-py', 'fuzzywuzzy', 'python-Levenshtein' are installed.")  # Added fuzzywuzzy
    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}") 
        # More detailed error hints 
        if "java" in str(e).lower():
            print("\nHint: Tabula-py requires Java. Please install Java and ensure it's in your system's PATH.")
            print("Download Java: https://www.oracle.com/java/technologies/downloads/")
        if args.debug:
            traceback.print_exc() # Print full traceback if debug is on 